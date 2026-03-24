"""Excel report writer for RawSourcesDetection results.

Produces an 8-sheet workbook:
  1. Summary              — run settings and stats
  2. Required Inputs      — every upstream source that must be supplied
  3. Formula Dependency   — all external formula references by level
  4. Missing Files        — referenced files not found on disk
  5. Matched Vectors      — hardcoded vectors with confirmed source
  6. Unmatched Vectors    — hardcoded vectors with no source found (+ context)
  7. Dynamic References   — INDIRECT / RTD / invisible live data feeds
  8. Stale Links          — phantom xl/externalLinks/ no longer used by formulas
  9. Scenarios            — Excel Scenario Manager entries (hidden inputs)
"""
from __future__ import annotations

from pathlib import Path

from .config import RSDConfig
from .models import DetectionResult


# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
_HDR_FILL    = "1565C0"   # dark blue — header background
_HDR_TEXT    = "FFFFFF"   # white — header text
_FOUND_FILL  = "C8E6C9"   # green — file found on disk
_MISSING_FILL = "FFCDD2"  # red/salmon — file not found
_MISS_ALT    = "FFEBEE"   # lighter red — alternating missing row
_EXACT_FILL  = "C8E6C9"   # green — exact match
_SUBSEQ_FILL = "DCEDC8"   # light green — exact subsequence
_APPROX_FILL = "BBDEFB"   # blue — approximate match
_UNMATCH_FILL = "FFF9C4"  # yellow — unmatched vector
_UNMATCH_ALT  = "FFFDE7"  # lighter yellow — alternating
_DB_FILL     = "E3F2FD"   # blue — database / ODBC / OLE DB
_FILE_FILL   = "E8F5E9"   # green — file reference
_PQ_FILL     = "FFF3E0"   # orange — Power Query
_OTHER_FILL  = "F3E5F5"   # purple — other categories
_DYN_FILL    = "FFE0B2"   # orange — dynamic / invisible reference
_DYN_ALT     = "FFF3E0"   # lighter orange — alternating
_STALE_FILL  = "E0E0E0"   # grey — stale/phantom link
_STALE_ALT   = "F5F5F5"   # lighter grey — alternating
_SCEN_FILL   = "E8EAF6"   # indigo tint — scenario
_SCEN_ALT    = "F3E5F5"   # lighter — alternating


# Map category → background colour for Required Inputs sheet
_CAT_FILL = {
    "database":   _DB_FILL,
    "powerquery": _PQ_FILL,
    "file":       _FILE_FILL,
    "formula":    _FILE_FILL,
}

# Categories to EXCLUDE from Required Inputs (not external data sources)
_SKIP_CATEGORIES = {"input", "metadata", "hardcoded"}

# Map match_type → background colour for Matched Vectors sheet
_MATCH_FILL = {
    "exact":             _EXACT_FILL,
    "exact_subsequence": _SUBSEQ_FILL,
    "approximate":       _APPROX_FILL,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fmt_sample(values: list[float]) -> str:
    """Format first 5 sample values compactly."""
    parts = []
    for v in values[:5]:
        if isinstance(v, float) and v == int(v) and abs(v) < 1e12:
            parts.append(str(int(v)))
        else:
            parts.append(f"{v:.4g}")
    return ", ".join(parts)


# ---------------------------------------------------------------------------
# Main writer
# ---------------------------------------------------------------------------

def write_report(
    result: DetectionResult,
    config: RSDConfig,
    model_path: Path,
    out_path: Path,
) -> None:
    """Write the full DetectionResult to an Excel workbook at *out_path*."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise ImportError(f"openpyxl required: {e}")

    wb = Workbook()
    wb.remove(wb.active)

    # ── Shared style factories ──────────────────────────────────────────────

    hdr_fill  = PatternFill("solid", fgColor=_HDR_FILL)
    hdr_font  = Font(bold=True, color=_HDR_TEXT, size=11)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    def pfill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def hdr(ws, headers: list[str], widths: list[int] | None = None) -> None:
        """Write header row with blue fill."""
        ws.row_dimensions[1].height = 30
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(1, ci, h)
            cell.fill  = hdr_fill
            cell.font  = hdr_font
            cell.alignment = hdr_align
            cell.border = border
        if widths:
            for ci, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w

    def cell(ws, row: int, col: int, value, fill=None,
             bold: bool = False, wrap: bool = False):
        """Write a data cell with consistent styling."""
        c = ws.cell(row, col, value)
        c.border = border
        c.font   = Font(bold=bold, size=10)
        c.alignment = Alignment(wrap_text=wrap, vertical="top")
        if fill:
            c.fill = fill
        return c

    def autowidth(ws, min_w: int = 10, max_w: int = 50) -> None:
        """Auto-size column widths, clamped to [min_w, max_w]."""
        for col_cells in ws.columns:
            best = max(
                (len(str(c.value)) for c in col_cells if c.value is not None),
                default=0,
            )
            ci = col_cells[0].column
            ws.column_dimensions[get_column_letter(ci)].width = max(min_w, min(max_w, best + 2))

    # ========================================================================
    # Sheet 1: Summary
    # ========================================================================
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 50

    ws["A1"] = "RawSourcesDetection Report"
    ws["A1"].font = Font(bold=True, size=14, color="1565C0")
    ws.merge_cells("A1:B1")

    n_found = sum(1 for n in result.source_nodes if n.found_on_disk and n.level > 0)

    # Count connections excluding internal/metadata noise
    visible_sources = [
        s for s in result.raw_sources
        if s.category not in _SKIP_CATEGORIES
    ]

    rows: list[tuple[str, object]] = [
        ("", ""),
        ("Model File", result.model_file),
        ("Sheets Traced", ", ".join(config.model_sheets) if config.model_sheets else "(all)"),
        ("Max Formula Levels", config.max_formula_levels),
        ("Match Mode", "Exact + Approximate" if config.approximate else "Exact only"),
        ("Decimal Places (Exact)", config.exact_decimal_places),
        ("Subsequence Matching", "Yes" if config.subsequence_matching else "No"),
        ("Min Vector Length", config.min_vector_length),
        ("", ""),
        ("── Formula Tracing ──", ""),
        ("Files Found on Disk", n_found),
        ("Files Missing (not on disk)", len(result.missing_files)),
        ("Formula Refs Found", len(result.formula_refs)),
        ("Formula Levels Reached", max((r.level for r in result.formula_refs), default=0)),
        ("", ""),
        ("── Vector Matching ──", ""),
        ("Vectors Matched", len(result.matched_vectors)),
        ("Vectors Unmatched", len(result.unmatched_vectors)),
        ("", ""),
        ("── Connection Harvesting ──", ""),
        ("Raw Sources Found", len(visible_sources)),
        ("  — Database / ODBC / OLE DB",
         sum(1 for s in visible_sources if s.category == "database")),
        ("  — Power Query",
         sum(1 for s in visible_sources if s.category == "powerquery")),
        ("  — File References",
         sum(1 for s in visible_sources if s.category == "file")),
        ("  — Other",
         sum(1 for s in visible_sources
             if s.category not in ("database", "powerquery", "file"))),
        ("", ""),
        ("── Extra Scanners ──", ""),
        ("Dynamic INDIRECT refs", len(result.dynamic_refs)),
        ("RTD (live feed) refs", len(result.rtd_refs)),
        ("Stale / phantom links", len(result.phantom_links)),
        ("XLSB files (limited scan)", len(result.xlsb_warnings)),
        ("Scenario Manager entries", len(result.scenarios)),
    ]

    lbl_font = Font(bold=True, size=10)
    val_font = Font(size=10)
    for ri, (label, value) in enumerate(rows, 2):
        ws.cell(ri, 1, label).font = lbl_font
        ws.cell(ri, 2, str(value) if value != "" else "").font = val_font

    # ========================================================================
    # Sheet 2: Required Inputs (pure dump — "shopping list")
    # ========================================================================
    ws = wb.create_sheet("Required Inputs")
    hdr(ws,
        ["Type", "Sub-Type", "Connection / Path / File",
         "Found On Disk", "Discovered In", "Location / Cell"],
        [14, 16, 52, 15, 26, 32])

    ri = 2
    seen_req: set[str] = set()

    # Part A: Excel files from formula tracing (deduplicated by filename)
    for ref in result.formula_refs:
        key = ref.target_file.lower()
        if key in seen_req:
            continue
        seen_req.add(key)
        f = pfill(_FOUND_FILL if ref.file_found else _MISSING_FILL)
        cell(ws, ri, 1, "Excel File", f)
        cell(ws, ri, 2, "formula_ref", f)
        cell(ws, ri, 3, ref.target_file, f)
        cell(ws, ri, 4, "Yes" if ref.file_found else "NO — MISSING", f, bold=not ref.file_found)
        cell(ws, ri, 5, ref.source_file, f)
        cell(ws, ri, 6, f"{ref.source_sheet}!{ref.source_cell}" if ref.source_cell else ref.source_sheet, f)
        ri += 1

    # Part B: ODBC / OLE DB / Power Query / other connections
    # Skip 'input', 'metadata' categories — not external data sources
    for src in result.raw_sources:
        if src.category in _SKIP_CATEGORIES:
            continue
        # Skip if same file already listed from formula_refs section
        # (avoid double-listing a .xlsx that appears as both formula ref and file connection)
        conn_lower = src.connection.lower()
        # Use connection string as dedup key for non-file types; filename for file types
        if src.category in ("file", "formula"):
            import os
            conn_key = os.path.basename(conn_lower)
            if conn_key in seen_req:
                continue
            seen_req.add(conn_key)
        else:
            key = f"{src.category}|{src.sub_type}|{conn_lower}"
            if key in seen_req:
                continue
            seen_req.add(key)

        f = pfill(_CAT_FILL.get(src.category, _OTHER_FILL))
        cell(ws, ri, 1, src.category, f)
        cell(ws, ri, 2, src.sub_type, f)
        cell(ws, ri, 3, src.connection, f, wrap=True)
        cell(ws, ri, 4, "N/A", f)
        cell(ws, ri, 5, src.source_file, f)
        cell(ws, ri, 6, src.location, f)
        ri += 1

    autowidth(ws)

    # ========================================================================
    # Sheet 3: Formula Dependency Tree
    # ========================================================================
    ws = wb.create_sheet("Formula Dependency Tree")
    hdr(ws,
        ["Level", "Source File", "Source Sheet", "Source Cell",
         "Target File", "Target Sheet", "Target Range",
         "Found on Disk", "Ref Origin", "Resolved Path"],
        [8, 24, 20, 12, 24, 20, 18, 14, 16, 42])

    for ri, ref in enumerate(result.formula_refs, 2):
        f = pfill(_FOUND_FILL if ref.file_found else _MISSING_FILL)
        cell(ws, ri, 1, ref.level, f)
        cell(ws, ri, 2, ref.source_file, f)
        cell(ws, ri, 3, ref.source_sheet, f)
        cell(ws, ri, 4, ref.source_cell, f)
        cell(ws, ri, 5, ref.target_file, f, bold=not ref.file_found)
        cell(ws, ri, 6, ref.target_sheet, f)
        cell(ws, ri, 7, ref.target_range, f)
        cell(ws, ri, 8, "Yes" if ref.file_found else "NO", f)
        cell(ws, ri, 9, getattr(ref, "ref_origin", "formula"), f)
        cell(ws, ri, 10, ref.resolved_path, f)

    autowidth(ws)

    # ========================================================================
    # Sheet 4: Missing Files
    # ========================================================================
    ws = wb.create_sheet("Missing Files")
    hdr(ws,
        ["Missing File", "Level", "Referenced By",
         "Sheets Needed", "Cells Referencing (first 10)",
         "Transitive Unknown"],
        [32, 8, 28, 32, 44, 20])

    for ri, mf in enumerate(result.missing_files, 2):
        f = pfill(_MISSING_FILL if ri % 2 == 0 else _MISS_ALT)
        cell(ws, ri, 1, mf.filename, f, bold=True)
        cell(ws, ri, 2, mf.level, f)
        cell(ws, ri, 3, mf.referenced_by, f)
        cell(ws, ri, 4, ", ".join(mf.sheets_needed), f, wrap=True)
        refs_text = "\n".join(mf.cells_referencing[:10])
        cell(ws, ri, 5, refs_text, f, wrap=True)
        transitive_note = (
            "YES — all dependencies of this file are invisible until it is supplied"
            if mf.transitive_unknown else "No"
        )
        cell(ws, ri, 6, transitive_note, f, wrap=True)
        if mf.cells_referencing:
            ws.row_dimensions[ri].height = min(15 * min(len(mf.cells_referencing), 10), 80)

    autowidth(ws)

    # ========================================================================
    # Sheet 5: Matched Vectors
    # ========================================================================
    ws = wb.create_sheet("Matched Vectors")
    hdr(ws,
        ["Model Sheet", "Model Range", "Length", "Model Sample (first 5)",
         "Match Type", "Similarity",
         "Upstream File", "Upstream Sheet", "Upstream Range", "Upstream Sample"],
        [18, 14, 8, 30, 18, 10, 28, 18, 16, 30])

    for ri, mv in enumerate(result.matched_vectors, 2):
        f = pfill(_MATCH_FILL.get(mv.match_type, _EXACT_FILL))
        cell(ws, ri, 1, mv.model_sheet, f)
        cell(ws, ri, 2, mv.model_range, f)
        cell(ws, ri, 3, mv.model_length, f)
        cell(ws, ri, 4, _fmt_sample(mv.model_sample), f)
        cell(ws, ri, 5, mv.match_type, f)
        cell(ws, ri, 6, round(mv.similarity, 6), f)
        cell(ws, ri, 7, mv.upstream_file, f)
        cell(ws, ri, 8, mv.upstream_sheet, f)
        cell(ws, ri, 9, mv.upstream_range, f)
        cell(ws, ri, 10, _fmt_sample(mv.upstream_sample), f)

    autowidth(ws)

    # ========================================================================
    # Sheet 6: Unmatched Vectors (with context)
    # ========================================================================
    ws = wb.create_sheet("Unmatched Vectors")
    hdr(ws,
        ["Model Sheet", "Model Range", "Length",
         "Column Header", "Row Label",
         "Sample Values (first 5)", "Action Required"],
        [18, 14, 8, 24, 24, 36, 48])

    for ri, uv in enumerate(result.unmatched_vectors, 2):
        f = pfill(_UNMATCH_FILL if ri % 2 == 0 else _UNMATCH_ALT)
        cell(ws, ri, 1, uv.model_sheet, f)
        cell(ws, ri, 2, uv.model_range, f)
        cell(ws, ri, 3, uv.model_length, f)
        cell(ws, ri, 4, getattr(uv, "column_header", ""), f)
        cell(ws, ri, 5, getattr(uv, "row_label", ""), f)
        cell(ws, ri, 6, _fmt_sample(uv.model_sample), f)
        cell(ws, ri, 7, "No source found — verify manually or supply upstream file", f)

    autowidth(ws)

    # ========================================================================
    # Sheet 7: Dynamic References
    # (INDIRECT with runtime-assembled filenames + RTD live COM feeds)
    # ========================================================================
    ws = wb.create_sheet("Dynamic References")

    has_dynamic = result.dynamic_refs or result.rtd_refs
    if not has_dynamic:
        ws["A1"] = "No dynamic references detected."
        ws["A1"].font = Font(italic=True, size=10)
    else:
        hdr(ws,
            ["Type", "Source File", "Source Sheet", "Source Cell",
             "ProgID / Detail", "Formula (truncated)", "Note"],
            [14, 26, 20, 12, 24, 50, 48])

        ri = 2
        for dr in result.dynamic_refs:
            f = pfill(_DYN_FILL if ri % 2 == 0 else _DYN_ALT)
            cell(ws, ri, 1, "INDIRECT (dynamic)", f, bold=True)
            cell(ws, ri, 2, dr.source_file, f)
            cell(ws, ri, 3, dr.source_sheet, f)
            cell(ws, ri, 4, dr.source_cell, f)
            cell(ws, ri, 5, "", f)
            cell(ws, ri, 6, dr.formula, f, wrap=True)
            cell(ws, ri, 7, dr.note, f, wrap=True)
            ri += 1
        for rr in result.rtd_refs:
            f = pfill(_DYN_FILL if ri % 2 == 0 else _DYN_ALT)
            cell(ws, ri, 1, "RTD (live feed)", f, bold=True)
            cell(ws, ri, 2, rr.source_file, f)
            cell(ws, ri, 3, rr.source_sheet, f)
            cell(ws, ri, 4, rr.source_cell, f)
            cell(ws, ri, 5, rr.prog_id, f)
            cell(ws, ri, 6, rr.formula, f, wrap=True)
            cell(ws, ri, 7, "Requires live COM server — not a file dependency", f, wrap=True)
            ri += 1

        autowidth(ws)

    # ========================================================================
    # Sheet 8: Stale Links
    # ========================================================================
    ws = wb.create_sheet("Stale Links")

    if not result.phantom_links:
        ws["A1"] = "No stale external links detected."
        ws["A1"].font = Font(italic=True, size=10)
    else:
        hdr(ws,
            ["Source File", "Stale Linked File", "Recommendation"],
            [32, 44, 52])

        for ri, pl in enumerate(result.phantom_links, 2):
            f = pfill(_STALE_FILL if ri % 2 == 0 else _STALE_ALT)
            cell(ws, ri, 1, pl.source_file, f)
            cell(ws, ri, 2, pl.stale_filename, f, bold=True)
            cell(ws, ri, 3,
                 "Remove via Excel → Data → Edit Links → Break Link. "
                 "File is registered but not referenced by any formula.", f, wrap=True)

        autowidth(ws)

    # ========================================================================
    # Sheet 9: Scenarios
    # ========================================================================
    ws = wb.create_sheet("Scenarios")

    if not result.scenarios:
        ws["A1"] = "No Excel Scenario Manager entries detected."
        ws["A1"].font = Font(italic=True, size=10)
    else:
        hdr(ws,
            ["Source File", "Sheet", "Scenario Name", "Input Cell", "Value"],
            [28, 20, 28, 14, 24])

        ri = 2
        for se in result.scenarios:
            base_f = pfill(_SCEN_FILL if ri % 2 == 0 else _SCEN_ALT)
            if not se.input_cells:
                cell(ws, ri, 1, se.source_file, base_f)
                cell(ws, ri, 2, se.sheet_name, base_f)
                cell(ws, ri, 3, se.scenario_name, base_f, bold=True)
                cell(ws, ri, 4, "(no cells)", base_f)
                cell(ws, ri, 5, "", base_f)
                ri += 1
            else:
                for idx, (cell_ref, value) in enumerate(se.input_cells):
                    f = pfill(_SCEN_FILL if ri % 2 == 0 else _SCEN_ALT)
                    cell(ws, ri, 1, se.source_file if idx == 0 else "", f)
                    cell(ws, ri, 2, se.sheet_name if idx == 0 else "", f)
                    cell(ws, ri, 3, se.scenario_name if idx == 0 else "", f, bold=(idx == 0))
                    cell(ws, ri, 4, cell_ref, f)
                    cell(ws, ri, 5, value, f)
                    ri += 1

        autowidth(ws)

    # ── Save ────────────────────────────────────────────────────────────────
    wb.save(str(out_path))
