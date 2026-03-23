#!/usr/bin/env python3
"""Convert an upstream tracing Excel report into a Mermaid flowchart.

Reads the Level 1, Level 2, ... sheets from an upstream_tracing_*.xlsx report
and produces a clean Mermaid diagram showing file → sheet → range connections
across all levels.

Usage:
    python trace_upstream_mermaid.py upstream_tracing_model.xlsx
    python trace_upstream_mermaid.py upstream_tracing_model.xlsx -o diagram.md
    python trace_upstream_mermaid.py upstream_tracing_model.xlsx --lr
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path


def _sanitize_id(text: str) -> str:
    """Turn arbitrary text into a safe Mermaid node ID."""
    return re.sub(r"[^A-Za-z0-9_]", "_", text)


def _short_name(path_or_name: str) -> str:
    """Extract just the filename from a full path or return as-is."""
    if not path_or_name:
        return "(unknown)"
    # Could be a full path or just a filename
    name = Path(path_or_name).name if ("/" in path_or_name or "\\" in path_or_name) else path_or_name
    return name or path_or_name


def _read_levels(report_path: Path) -> list[dict]:
    """Read all Level N sheets and return a flat list of edge dicts."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Error: openpyxl is required. Install with: pip install openpyxl",
              file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(str(report_path), read_only=True, data_only=True)

    # Find Level sheets
    level_sheets = sorted(
        [s for s in wb.sheetnames if re.match(r"Level \d+", s)],
        key=lambda s: int(s.split()[-1]),
    )

    if not level_sheets:
        print("Error: No 'Level N' sheets found in the report.", file=sys.stderr)
        wb.close()
        sys.exit(1)

    edges: list[dict] = []

    for sheet_name in level_sheets:
        level = int(sheet_name.split()[-1])
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or not row[0]:
                continue
            # Columns: A=Source File, B=Source Sheet, C=Source Cell,
            #          D=Formula, E=Target File, F=Target Sheet,
            #          G=Target Range, H=Target Path, I=File Found,
            #          J=Resolved Path, K=Precedent Chain
            edges.append({
                "level": level,
                "src_file": str(row[0] or ""),
                "src_sheet": str(row[1] or ""),
                "src_range": str(row[2] or ""),
                "tgt_file": str(row[4] or ""),
                "tgt_sheet": str(row[5] or ""),
                "tgt_range": str(row[6] or ""),
                "file_found": str(row[8] or "").strip().lower() == "yes",
            })

    wb.close()
    return edges


def _read_config_model(report_path: Path) -> str | None:
    """Read the model filename from the Config sheet."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    wb = load_workbook(str(report_path), read_only=True, data_only=True)
    if "Config" not in wb.sheetnames:
        wb.close()
        return None

    ws = wb["Config"]
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if row and row[0] and str(row[0]).strip() == "Model file":
            wb.close()
            return _short_name(str(row[1] or ""))
    wb.close()
    return None


def _build_mermaid(
    edges: list[dict],
    model_name: str | None,
    direction: str = "TB",
) -> str:
    """Build the Mermaid flowchart string."""

    # --- Aggregate edges: group source cells pointing to the same target ---
    # Key: (src_file, src_sheet, tgt_file, tgt_sheet, tgt_range, file_found)
    # Value: set of source cells
    agg: dict[tuple, set[str]] = {}
    for e in edges:
        key = (
            e["src_file"], e["src_sheet"],
            e["tgt_file"], e["tgt_sheet"], e["tgt_range"],
            e["file_found"], e["level"],
        )
        agg.setdefault(key, set()).add(e["src_range"])

    # --- Collect unique files, sheets, and build edges ---
    # Node IDs: file level -> f_<sanitized>, sheet level -> s_<file>_<sheet>
    file_nodes: dict[str, bool] = {}  # filename -> found (True if ever found)
    sheet_nodes: set[tuple[str, str]] = set()  # (file, sheet)

    link_lines: list[str] = []

    for (src_file, src_sheet, tgt_file, tgt_sheet, tgt_range, found, level), src_cells in agg.items():
        sf = _short_name(src_file)
        tf = _short_name(tgt_file)

        # Track files
        if sf not in file_nodes:
            file_nodes[sf] = True  # source files are always on disk
        if tf not in file_nodes:
            file_nodes[tf] = found
        elif found:
            file_nodes[tf] = True  # upgrade to found if any ref is found

        # Track sheets
        sheet_nodes.add((sf, src_sheet))
        sheet_nodes.add((tf, tgt_sheet))

        # Build label: aggregate source cells, show target range
        src_list = sorted(src_cells)
        if len(src_list) <= 3:
            src_label = ", ".join(src_list)
        else:
            src_label = f"{src_list[0]}, ... ({len(src_list)} cells)"

        tgt_label = tgt_range if tgt_range else "?"

        src_sid = _sanitize_id(f"s_{sf}_{src_sheet}")
        tgt_sid = _sanitize_id(f"s_{tf}_{tgt_sheet}")

        link_lines.append(
            f"    {src_sid} -->|\"'{src_label} → {tgt_label}'\"| {tgt_sid}"
        )

    # --- Build the output ---
    lines = [f"flowchart {direction}"]

    # Add model file as a special node if known
    if model_name and model_name not in file_nodes:
        file_nodes[model_name] = True

    # Group by file using subgraphs
    # Sort: model file first (if known), then alphabetically
    def file_sort_key(f: str) -> tuple[int, str]:
        if model_name and f == model_name:
            return (0, f)
        return (1, f)

    for fname in sorted(file_nodes.keys(), key=file_sort_key):
        fid = _sanitize_id(f"f_{fname}")
        found = file_nodes[fname]

        # Determine file's sheets
        file_sheets = sorted(s for f, s in sheet_nodes if f == fname)
        if not file_sheets:
            continue

        # Style marker for missing files
        style_class = "found" if found else "missing"

        lines.append(f"    subgraph {fid}[\"{fname}\"]")
        for sh in file_sheets:
            sid = _sanitize_id(f"s_{fname}_{sh}")
            lines.append(f"        {sid}[\"{sh}\"]")
        lines.append("    end")
        if not found:
            lines.append(f"    class {fid} missing")

    # Blank line before links
    lines.append("")

    # Deduplicate links
    seen_links: set[str] = set()
    for link in link_lines:
        if link not in seen_links:
            seen_links.add(link)
            lines.append(link)

    # Style definitions
    lines.append("")
    lines.append("    classDef missing fill:#FFCDD2,stroke:#C62828,stroke-width:2px,color:#B71C1C")
    lines.append("    classDef found fill:#C8E6C9,stroke:#2E7D32,stroke-width:1px")

    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(
        description="Convert an upstream tracing Excel report into a Mermaid flowchart",
    )
    parser.add_argument(
        "report",
        help="Path to the upstream_tracing_*.xlsx report file",
    )
    parser.add_argument(
        "-o", "--output",
        help="Output file path (default: <report_stem>_mermaid.md)",
    )
    parser.add_argument(
        "--lr",
        action="store_true",
        help="Use left-to-right layout instead of top-to-bottom",
    )
    args = parser.parse_args()

    report_path = Path(args.report).resolve()
    if not report_path.exists():
        print(f"Error: File not found: {report_path}", file=sys.stderr)
        sys.exit(1)

    # Read data
    edges = _read_levels(report_path)
    model_name = _read_config_model(report_path)

    direction = "LR" if args.lr else "TB"
    mermaid = _build_mermaid(edges, model_name, direction=direction)

    # Output
    if args.output:
        out_path = Path(args.output).resolve()
    else:
        out_path = report_path.with_name(report_path.stem + "_mermaid.md")

    content = f"```mermaid\n{mermaid}\n```\n"
    out_path.write_text(content, encoding="utf-8")

    print(f"Mermaid diagram: {out_path}")
    print(f"  Levels: {len(set(e['level'] for e in edges))}")
    print(f"  Edges:  {len(edges)}")
    print(f"  Files:  {len(set(e['src_file'] for e in edges) | set(e['tgt_file'] for e in edges))}")


if __name__ == "__main__":
    main()
