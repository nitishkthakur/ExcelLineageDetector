"""Tests for the BusinessContract pipeline."""
from __future__ import annotations

import json
import sys
import zipfile
from pathlib import Path

import pytest

# Ensure project root is on path
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from BusinessContract.mcp_server.streaming import (
    CellInfo,
    _load_shared_strings,
    col_letter_to_index,
    get_named_ranges,
    get_sheet_summary,
    index_to_col_letter,
    list_sheets,
    parse_ref,
    read_cell_neighborhood,
    stream_sheet_cells,
)
from BusinessContract.pipeline.config import ContractConfig
from BusinessContract.pipeline.formula_converter import batch_convert, excel_to_sql
from BusinessContract.pipeline.models import (
    BusinessContract,
    ContractVariable,
    DependencyEdge,
    TransformationStep,
)

FIXTURE_DIR = Path(__file__).parent / "data"
MODEL_PATH = FIXTURE_DIR / "model.xlsx"
UPSTREAM_A = FIXTURE_DIR / "upstream_a.xlsx"
UPSTREAM_B = FIXTURE_DIR / "upstream_b.xlsx"


# ── Helpers ──────────────────────────────────────────────────────

@pytest.fixture(autouse=True)
def ensure_fixtures():
    """Ensure fixture files exist."""
    if not MODEL_PATH.exists():
        from BusinessContract.fixtures.generate_fixtures import generate_all
        generate_all()
    assert MODEL_PATH.exists()
    assert UPSTREAM_A.exists()
    assert UPSTREAM_B.exists()


# ── Streaming Reader Tests ───────────────────────────────────────

class TestStreamingReader:
    def test_col_letter_to_index(self):
        assert col_letter_to_index("A") == 1
        assert col_letter_to_index("B") == 2
        assert col_letter_to_index("Z") == 26
        assert col_letter_to_index("AA") == 27
        assert col_letter_to_index("AZ") == 52

    def test_index_to_col_letter(self):
        assert index_to_col_letter(1) == "A"
        assert index_to_col_letter(26) == "Z"
        assert index_to_col_letter(27) == "AA"

    def test_parse_ref(self):
        assert parse_ref("B3") == (3, 2)
        assert parse_ref("A1") == (1, 1)
        assert parse_ref("AA100") == (100, 27)

    def test_list_sheets(self):
        with zipfile.ZipFile(str(MODEL_PATH), "r") as zf:
            sheets = list_sheets(zf)
        assert "Inputs" in sheets
        assert "Summary" in sheets
        assert len(sheets) == 4

    def test_stream_cells(self):
        with zipfile.ZipFile(str(MODEL_PATH), "r") as zf:
            ss = _load_shared_strings(zf)
            cells = list(stream_sheet_cells(zf, "Inputs", ss))
        assert len(cells) > 0
        refs = {c.ref for c in cells}
        assert "B2" in refs or "A1" in refs

    def test_read_neighborhood(self):
        with zipfile.ZipFile(str(MODEL_PATH), "r") as zf:
            ss = _load_shared_strings(zf)
            cells = read_cell_neighborhood(zf, "Inputs", "B2", radius=2, shared_strings=ss)
        assert len(cells) > 0

    def test_get_sheet_summary(self):
        with zipfile.ZipFile(str(MODEL_PATH), "r") as zf:
            ss = _load_shared_strings(zf)
            summary = get_sheet_summary(zf, "Inputs", ss)
        assert summary["sheet_name"] == "Inputs"
        assert summary["max_row"] > 0
        assert summary["max_col"] > 0
        assert isinstance(summary["headers"], list)

    def test_get_named_ranges(self):
        with zipfile.ZipFile(str(MODEL_PATH), "r") as zf:
            names = get_named_ranges(zf)
        assert len(names) >= 3  # DiscountRate, TerminalGrowth, TaxRate
        name_set = {n["name"] for n in names}
        assert "DiscountRate" in name_set

    def test_upstream_b_named_ranges(self):
        with zipfile.ZipFile(str(UPSTREAM_B), "r") as zf:
            names = get_named_ranges(zf)
        name_set = {n["name"] for n in names}
        assert "PriceHistory" in name_set
        assert "VolumeHistory" in name_set


# ── Formula Converter Tests ──────────────────────────────────────

class TestFormulaConverter:
    def test_basic_sum(self):
        assert "SUM(" in excel_to_sql("SUM(B2:B13)")

    def test_average_to_avg(self):
        assert "AVG(" in excel_to_sql("AVERAGE(B2:B13)")

    def test_if_to_case(self):
        result = excel_to_sql("IF(A1>0,B1,C1)")
        assert "CASE WHEN" in result

    def test_variable_substitution(self):
        var_names = {"Inputs!B2:B13": "revenue"}
        result = excel_to_sql("SUM(Inputs!B2:B13)", var_names)
        assert "revenue" in result

    def test_operator_replacement(self):
        result = excel_to_sql("A1<>B1")
        assert "!=" in result

    def test_concatenation(self):
        result = excel_to_sql('A1&B1')
        assert "||" in result

    def test_batch_convert(self):
        formulas = [
            {"formula": "SUM(A1:A10)", "cell": "B1"},
            {"formula": "AVERAGE(B1:B10)", "cell": "C1"},
        ]
        result = batch_convert(formulas)
        assert all("sql_formula" in f for f in result)


# ── Data Models Tests ────────────────────────────────────────────

class TestModels:
    def test_contract_variable_from_vector(self):
        var = ContractVariable.from_vector(
            sheet="Sheet1",
            cell_range="B2:B13",
            direction="column",
            length=12,
            values=[1.0, 2.0, 3.0, 4.0, 5.0, 6.0],
        )
        assert var.id
        assert var.sheet == "Sheet1"
        assert var.length == 12
        assert len(var.sample_values) == 5  # truncated to 5

    def test_transformation_step_make(self):
        tx = TransformationStep.make(
            output_id="out1",
            input_ids=["in1", "in2"],
            excel_formula="=SUM(A1:A10)",
            sql_formula="SUM(col_a)",
            sheet="Calc",
            cell_range="B1",
        )
        assert tx.id
        assert tx.output_variable_id == "out1"
        assert len(tx.input_variable_ids) == 2

    def test_business_contract_structure(self):
        bc = BusinessContract(
            model_file="test.xlsx",
            output_sheets=["Summary"],
            variables=[],
            transformations=[],
            edges=[],
        )
        assert bc.model_file == "test.xlsx"


# ── Config Tests ─────────────────────────────────────────────────

class TestConfig:
    def test_config_defaults(self):
        cfg = ContractConfig(
            model_path=Path("model.xlsx"),
            output_sheets=["Summary"],
        )
        assert cfg.llm_model == "claude-haiku-4-5-20251001"
        assert cfg.llm_batch_size == 20
        assert cfg.min_vector_length == 3

    def test_config_from_file(self, tmp_path):
        config_data = {
            "model_path": "model.xlsx",
            "output_sheets": ["Summary"],
            "llm_model": "claude-haiku-4-5-20251001",
        }
        cfg_file = tmp_path / "config.json"
        cfg_file.write_text(json.dumps(config_data))
        cfg = ContractConfig.from_file(cfg_file)
        assert cfg.model_path == Path("model.xlsx")
        assert cfg.output_sheets == ["Summary"]


# ── Scanner Tests ────────────────────────────────────────────────

class TestScanner:
    def test_scan_model_basic(self):
        from BusinessContract.pipeline.scanner import scan_model
        cfg = ContractConfig(
            model_path=MODEL_PATH,
            output_sheets=["Summary"],
        )
        contract = scan_model(cfg)
        assert len(contract.variables) > 0
        assert len(contract.connections) > 0

    def test_scan_finds_formulas(self):
        from BusinessContract.pipeline.scanner import _scan_formulas
        formulas = _scan_formulas(MODEL_PATH, ["Calculations"])
        assert len(formulas) > 0
        assert all("formula" in f for f in formulas)

    def test_scan_finds_hardcoded_vectors(self):
        from BusinessContract.pipeline.scanner import _scan_hardcoded_vectors
        vars = _scan_hardcoded_vectors(MODEL_PATH, ["Inputs"], min_length=3)
        assert len(vars) > 0
        assert all(v.source_type == "hardcoded" for v in vars)

    def test_output_sheets_marked(self):
        from BusinessContract.pipeline.scanner import scan_model
        cfg = ContractConfig(
            model_path=MODEL_PATH,
            output_sheets=["Summary"],
        )
        contract = scan_model(cfg)
        summary_vars = [v for v in contract.variables if v.sheet == "Summary"]
        for v in summary_vars:
            assert v.variable_type == "output"


# ── Graph Builder Tests ──────────────────────────────────────────

class TestGraphBuilder:
    def test_enrich_upstream(self):
        from BusinessContract.pipeline.graph_builder import enrich_upstream
        from BusinessContract.pipeline.scanner import scan_model
        cfg = ContractConfig(
            model_path=MODEL_PATH,
            output_sheets=["Summary"],
            upstream_dir=FIXTURE_DIR,
        )
        contract = scan_model(cfg)
        contract = enrich_upstream(contract, cfg)
        upstream_vars = [v for v in contract.variables if v.upstream_source]
        assert len(upstream_vars) > 0


# ── Contract Writer Tests ────────────────────────────────────────

class TestContractWriter:
    def test_write_contract(self, tmp_path):
        from BusinessContract.pipeline.contract_writer import write_contract
        from BusinessContract.pipeline.scanner import scan_model
        cfg = ContractConfig(
            model_path=MODEL_PATH,
            output_sheets=["Summary"],
        )
        contract = scan_model(cfg)
        # Assign fallback names
        for v in contract.variables:
            v.business_name = f"{v.sheet}_{v.cell_range}".lower()

        out = tmp_path / "test_contract.xlsx"
        write_contract(contract, out)
        assert out.exists()
        assert out.stat().st_size > 0

        # Verify sheet structure
        import openpyxl
        wb = openpyxl.load_workbook(str(out))
        assert "Summary" in wb.sheetnames
        assert "Variables" in wb.sheetnames
        assert "Transformations" in wb.sheetnames
        assert "Dependencies" in wb.sheetnames
        wb.close()


# ── Mermaid Generator Tests ──────────────────────────────────────

class TestMermaidGenerator:
    def _make_contract(self):
        from BusinessContract.pipeline.scanner import scan_model
        cfg = ContractConfig(
            model_path=MODEL_PATH,
            output_sheets=["Summary"],
        )
        contract = scan_model(cfg)
        for v in contract.variables:
            v.business_name = f"{v.sheet}_{v.cell_range}".lower()
        return contract

    def test_source_level_diagram(self):
        from BusinessContract.mermaid.generator import generate_source_level
        contract = self._make_contract()
        diagram = generate_source_level(contract)
        assert "graph LR" in diagram
        assert "classDef" in diagram

    def test_variable_level_diagram(self):
        from BusinessContract.mermaid.generator import generate_variable_level
        contract = self._make_contract()
        diagram = generate_variable_level(contract)
        assert "graph TD" in diagram
        assert "subgraph Inputs" in diagram

    def test_write_mermaid(self, tmp_path):
        from BusinessContract.mermaid.generator import write_mermaid
        contract = self._make_contract()
        src, var = write_mermaid(contract, tmp_path, "test_model")
        assert src.exists()
        assert var.exists()
        assert "```mermaid" in src.read_text()


# ── Python Refactor Generator Tests ──────────────────────────────

class TestRefactorGenerator:
    def test_generate_python(self, tmp_path):
        from BusinessContract.pipeline.scanner import scan_model
        from BusinessContract.refactor.generator import generate_python
        cfg = ContractConfig(
            model_path=MODEL_PATH,
            output_sheets=["Summary"],
        )
        contract = scan_model(cfg)
        for v in contract.variables:
            v.business_name = f"{v.sheet}_{v.cell_range}".lower()

        py_path = tmp_path / "engine.py"
        generate_python(contract, py_path)
        assert py_path.exists()

        content = py_path.read_text()
        assert "import numpy" in content
        assert "def compute_all" in content
        assert "if __name__" in content


# ── End-to-End Pipeline Test ─────────────────────────────────────

class TestEndToEnd:
    def test_full_pipeline_skip_llm(self, tmp_path):
        from BusinessContract.pipeline.run import generate_contract
        cfg = ContractConfig(
            model_path=MODEL_PATH,
            output_sheets=["Summary"],
            upstream_dir=FIXTURE_DIR,
            out_dir=tmp_path,
        )
        contract = generate_contract(cfg, skip_llm=True)

        assert len(contract.variables) > 0
        assert len(contract.transformations) > 0
        assert len(contract.connections) > 0

        # Check all outputs were written
        assert (tmp_path / "business_contract_model.xlsx").exists()
        assert (tmp_path / "model_source_flow.md").exists()
        assert (tmp_path / "model_variable_flow.md").exists()
        assert (tmp_path / "calculation_engine_model.py").exists()

    def test_variable_types_assigned(self):
        from BusinessContract.pipeline.scanner import scan_model
        cfg = ContractConfig(
            model_path=MODEL_PATH,
            output_sheets=["Summary"],
        )
        contract = scan_model(cfg)
        types = {v.variable_type for v in contract.variables}
        assert "input" in types
        assert "output" in types or "intermediate" in types


# ── Gap-fix regression tests ─────────────────────────────────────

class TestParseRefAbsolute:
    """Fix: parse_ref must handle $B$3 style refs."""
    def test_dollar_sign_stripped(self):
        assert parse_ref("$B$3") == (3, 2)

    def test_mixed_absolute(self):
        assert parse_ref("$AA$100") == (100, 27)


class TestRangeOverlap:
    """Fix: proper rectangle-based range overlap."""
    def test_single_cell_in_range(self):
        from BusinessContract.pipeline.scanner import _ranges_overlap
        assert _ranges_overlap("B3", "B2:B13") is True

    def test_single_cell_outside_range(self):
        from BusinessContract.pipeline.scanner import _ranges_overlap
        assert _ranges_overlap("Z1", "B2:B13") is False

    def test_range_vs_range_overlap(self):
        from BusinessContract.pipeline.scanner import _ranges_overlap
        assert _ranges_overlap("B2:B5", "B3:B10") is True

    def test_range_no_overlap(self):
        from BusinessContract.pipeline.scanner import _ranges_overlap
        assert _ranges_overlap("A1:A5", "C1:C5") is False

    def test_dollar_signs_ignored(self):
        from BusinessContract.pipeline.scanner import _ranges_overlap
        assert _ranges_overlap("$B$2", "B2:B13") is True

    def test_graph_builder_ranges_overlap(self):
        from BusinessContract.pipeline.graph_builder import _ranges_overlap
        assert _ranges_overlap("$A$1:$A$10", "A1:A10") is True


class TestScalarDetection:
    """Fix: formula scalars must be detected (not just vectors)."""
    def test_scalars_found(self):
        from BusinessContract.pipeline.scanner import scan_model
        cfg = ContractConfig(
            model_path=MODEL_PATH,
            output_sheets=["Summary"],
        )
        contract = scan_model(cfg)
        directions = {v.direction for v in contract.variables}
        assert "scalar" in directions, f"No scalar variables found; directions: {directions}"


class TestNamedRangeDetection:
    """Fix: named ranges must produce variables."""
    def test_named_ranges_scanned(self):
        from BusinessContract.pipeline.scanner import _scan_named_ranges
        named = _scan_named_ranges(MODEL_PATH)
        names = {v.business_name for v in named}
        assert "DiscountRate" in names
        assert "TaxRate" in names


class TestNestedIfConversion:
    """Fix: IF with nested function calls must not break."""
    def test_nested_sumif(self):
        result = excel_to_sql("IF(SUMIF(A:A,\">0\",B:B)>100,C1,D1)")
        assert "CASE WHEN" in result
        assert "THEN" in result
        assert "ELSE" in result
        assert "END" in result

    def test_simple_if(self):
        result = excel_to_sql("IF(A1>0,B1,C1)")
        assert "CASE WHEN" in result
        assert "END" in result


class TestConfigValidation:
    """Fix: config must validate fields."""
    def test_empty_output_sheets_rejected(self):
        with pytest.raises(ValueError, match="output_sheets"):
            ContractConfig(model_path=Path("x.xlsx"), output_sheets=[])

    def test_bad_similarity_rejected(self):
        with pytest.raises(ValueError, match="min_similarity"):
            ContractConfig(model_path=Path("x.xlsx"), output_sheets=["S"], min_similarity=1.5)

    def test_missing_key_in_file(self, tmp_path):
        cfg_file = tmp_path / "bad.json"
        cfg_file.write_text('{"model_path": "x.xlsx"}')
        with pytest.raises(ValueError, match="output_sheets"):
            ContractConfig.from_file(cfg_file)


class TestEdgeDedup:
    """Fix: edges must be deduplicated."""
    def test_no_duplicate_edges(self):
        from BusinessContract.pipeline.scanner import scan_model
        cfg = ContractConfig(
            model_path=MODEL_PATH,
            output_sheets=["Summary"],
        )
        contract = scan_model(cfg)
        edge_keys = [(e.source_id, e.target_id) for e in contract.edges]
        assert len(edge_keys) == len(set(edge_keys)), "Duplicate edges found"
