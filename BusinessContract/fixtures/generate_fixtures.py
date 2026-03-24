"""Generate complex multi-file test fixtures for the Business Contract pipeline.

Creates three interconnected Excel files:
  model.xlsx      -- 4 sheets, formulas, hardcoded vectors, OLEDB connection, external links
  upstream_a.xlsx -- 2 sheets, named ranges, a table, formulas referencing upstream_b
  upstream_b.xlsx -- 1 sheet, raw data tables (terminal node)

The hardcoded vectors in model.xlsx are exact sub-sequences of data in upstream_a.xlsx
so that nearest-neighbour matching finds them.
"""
from __future__ import annotations

import re
import zipfile
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Font
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo


# ---------------------------------------------------------------------------
# Shared data — vectors that will be planted in upstream_b → upstream_a → model
# ---------------------------------------------------------------------------

# Raw price history (40 values) — lives in upstream_b
PRICES = [float(x) for x in np.round(np.cumsum(np.random.default_rng(42).normal(0, 2, 40)) + 100, 2)]

# Revenue vector (12 quarterly values) — subset of PRICES[0:12], lives in upstream_a
REVENUE_Q = PRICES[:12]

# Cost vector (12 values)
COSTS_Q = [round(r * 0.6 + np.random.default_rng(99).normal(0, 1), 2) for r in REVENUE_Q]

# Growth rates (12 values)
GROWTH = [round((REVENUE_Q[i] - REVENUE_Q[i - 1]) / abs(REVENUE_Q[i - 1]) if i > 0 else 0, 4) for i in range(12)]

# Volume history (40 values) — lives in upstream_b
VOLUMES = [int(x) for x in np.round(np.cumsum(np.random.default_rng(7).normal(0, 50, 40)) + 5000)]


# ---------------------------------------------------------------------------
# upstream_b.xlsx — raw data, terminal node
# ---------------------------------------------------------------------------

def _generate_upstream_b(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RawData"

    # Headers
    ws["A1"] = "Date"
    ws["B1"] = "Price"
    ws["C1"] = "Volume"
    ws["D1"] = "Sector"
    for c in ["A1", "B1", "C1", "D1"]:
        ws[c].font = Font(bold=True)

    # Data — 40 rows
    for i in range(40):
        ws.cell(row=i + 2, column=1, value=f"2024-Q{(i % 4) + 1}")
        ws.cell(row=i + 2, column=2, value=PRICES[i])
        ws.cell(row=i + 2, column=3, value=VOLUMES[i])
        ws.cell(row=i + 2, column=4, value="Technology")

    # Define a named range for prices
    wb.defined_names.add(DefinedName("PriceHistory", attr_text="RawData!$B$2:$B$41"))
    wb.defined_names.add(DefinedName("VolumeHistory", attr_text="RawData!$C$2:$C$41"))

    wb.save(str(path))


# ---------------------------------------------------------------------------
# upstream_a.xlsx — intermediate, references upstream_b
# ---------------------------------------------------------------------------

def _generate_upstream_a(path: Path, upstream_b_name: str = "upstream_b.xlsx") -> None:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Data ─────────────────────────────────────────────
    ws_data = wb.active
    ws_data.title = "Data"
    ws_data["A1"] = "Quarter"
    ws_data["B1"] = "Revenue"
    ws_data["C1"] = "Cost"
    ws_data["D1"] = "Source"
    for c in ["A1", "B1", "C1", "D1"]:
        ws_data[c].font = Font(bold=True)

    for i in range(12):
        ws_data.cell(row=i + 2, column=1, value=f"Q{i + 1}")
        ws_data.cell(row=i + 2, column=2, value=REVENUE_Q[i])
        ws_data.cell(row=i + 2, column=3, value=COSTS_Q[i])
        ws_data.cell(row=i + 2, column=4, value="Bloomberg")

    # Named range for revenue
    wb.defined_names.add(DefinedName("QuarterlyRevenue", attr_text="Data!$B$2:$B$13"))

    # Table on the data
    tab = Table(displayName="tbl_quarterly", ref="A1:D13")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9")
    ws_data.add_table(tab)

    # ── Sheet 2: Lookup ───────────────────────────────────────────
    ws_lookup = wb.create_sheet("Lookup")
    ws_lookup["A1"] = "Metric"
    ws_lookup["B1"] = "Value"
    ws_lookup["A2"] = "Discount Rate"
    ws_lookup["B2"] = 0.08
    ws_lookup["A3"] = "Terminal Growth"
    ws_lookup["B3"] = 0.025
    ws_lookup["A4"] = "Tax Rate"
    ws_lookup["B4"] = 0.21
    wb.defined_names.add(DefinedName("DiscountRate", attr_text="Lookup!$B$2"))
    wb.defined_names.add(DefinedName("TaxRate", attr_text="Lookup!$B$4"))

    wb.save(str(path))

    # ── Inject external links to upstream_b via XML ───────────────
    _inject_external_links(path, upstream_b_name, [
        ("Lookup", "C2", f"'[{upstream_b_name}]RawData'!B2"),
        ("Lookup", "C3", f"'[{upstream_b_name}]RawData'!B3"),
    ])


# ---------------------------------------------------------------------------
# model.xlsx — the main model file
# ---------------------------------------------------------------------------

def _generate_model(path: Path, upstream_a_name: str = "upstream_a.xlsx") -> None:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Inputs ───────────────────────────────────────────
    ws_in = wb.active
    ws_in.title = "Inputs"
    ws_in["A1"] = "Quarterly Revenue (Source: Bloomberg)"
    ws_in["A1"].font = Font(bold=True)
    for i, v in enumerate(REVENUE_Q):
        ws_in.cell(row=2, column=i + 2, value=v)  # row vector B2:M2

    ws_in["A3"] = "Quarterly Costs"
    ws_in["A3"].font = Font(bold=True)
    for i, v in enumerate(COSTS_Q):
        ws_in.cell(row=4, column=i + 2, value=v)  # row vector B4:M4

    ws_in["A6"] = "Growth Rates"
    ws_in["A6"].font = Font(bold=True)
    for i, v in enumerate(GROWTH):
        ws_in.cell(row=7, column=i + 2, value=v)  # row vector B7:M7

    ws_in["A9"] = "Volume Data"
    ws_in["A9"].font = Font(bold=True)
    for i in range(12):
        ws_in.cell(row=10, column=i + 2, value=float(VOLUMES[i]))

    # ── Sheet 2: Assumptions ──────────────────────────────────────
    ws_assume = wb.create_sheet("Assumptions")
    ws_assume["A1"] = "Parameter"
    ws_assume["B1"] = "Value"
    ws_assume["A1"].font = Font(bold=True)
    ws_assume["B1"].font = Font(bold=True)

    ws_assume["A2"] = "Discount Rate"
    ws_assume["B2"] = 0.08
    ws_assume["A3"] = "Terminal Growth Rate"
    ws_assume["B3"] = 0.025
    ws_assume["A4"] = "Tax Rate"
    ws_assume["B4"] = 0.21
    ws_assume["A5"] = "Inflation Rate"
    ws_assume["B5"] = 0.03

    wb.defined_names.add(DefinedName("DiscountRate", attr_text="Assumptions!$B$2"))
    wb.defined_names.add(DefinedName("TerminalGrowth", attr_text="Assumptions!$B$3"))
    wb.defined_names.add(DefinedName("TaxRate", attr_text="Assumptions!$B$4"))

    # ── Sheet 3: Calculations ─────────────────────────────────────
    ws_calc = wb.create_sheet("Calculations")
    ws_calc["A1"] = "Metric"
    ws_calc["A1"].font = Font(bold=True)

    # Row 2: Gross Profit = Revenue - Cost (formulas)
    ws_calc["A2"] = "Gross Profit"
    for i in range(12):
        col = chr(66 + i)  # B..M
        ws_calc[f"{col}2"] = f"=Inputs!{col}2-Inputs!{col}4"

    # Row 3: Tax Amount = Gross Profit * TaxRate
    ws_calc["A3"] = "Tax Amount"
    for i in range(12):
        col = chr(66 + i)
        ws_calc[f"{col}3"] = f"={col}2*Assumptions!$B$4"

    # Row 4: Net Income = Gross Profit - Tax
    ws_calc["A4"] = "Net Income"
    for i in range(12):
        col = chr(66 + i)
        ws_calc[f"{col}4"] = f"={col}2-{col}3"

    # Row 5: Discounted CF = Net Income / (1+DiscountRate)^period
    ws_calc["A5"] = "Discounted Cash Flow"
    for i in range(12):
        col = chr(66 + i)
        ws_calc[f"{col}5"] = f"={col}4/(1+Assumptions!$B$2)^{i+1}"

    # Row 6: Cumulative DCF
    ws_calc["A6"] = "Cumulative DCF"
    ws_calc["B6"] = "=B5"
    for i in range(1, 12):
        col = chr(66 + i)
        prev = chr(65 + i)
        ws_calc[f"{col}6"] = f"={prev}6+{col}5"

    # ── Sheet 4: Summary (output sheet) ───────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = "Summary Output"
    ws_sum["A1"].font = Font(bold=True, size=14)

    ws_sum["A3"] = "Total Revenue"
    ws_sum["B3"] = "=SUM(Inputs!B2:M2)"
    ws_sum["A4"] = "Total Costs"
    ws_sum["B4"] = "=SUM(Inputs!B4:M4)"
    ws_sum["A5"] = "Total Gross Profit"
    ws_sum["B5"] = "=SUM(Calculations!B2:M2)"
    ws_sum["A6"] = "Total Net Income"
    ws_sum["B6"] = "=SUM(Calculations!B4:M4)"
    ws_sum["A7"] = "NPV (Sum of DCF)"
    ws_sum["B7"] = "=SUM(Calculations!B5:M5)"
    ws_sum["A8"] = "Terminal Value"
    ws_sum["B8"] = "=Calculations!M4*(1+Assumptions!B3)/(Assumptions!B2-Assumptions!B3)"
    ws_sum["A9"] = "Enterprise Value"
    ws_sum["B9"] = "=B7+B8"
    ws_sum["A10"] = "Discount Rate Used"
    ws_sum["B10"] = "=Assumptions!B2"
    ws_sum["A11"] = "Avg Growth Rate"
    ws_sum["B11"] = "=AVERAGE(Inputs!B7:M7)"

    wb.save(str(path))

    # ── Inject OLEDB connection + external links via XML ──────────
    _inject_oledb_connection(path)
    _inject_external_links(path, upstream_a_name, [
        ("Calculations", "N2", f"'[{upstream_a_name}]Data'!B2"),
        ("Calculations", "N3", f"'[{upstream_a_name}]'!QuarterlyRevenue"),
    ])


# ---------------------------------------------------------------------------
# XML injection helpers
# ---------------------------------------------------------------------------

def _inject_oledb_connection(xlsx_path: Path) -> None:
    """Inject an OLEDB connection into xl/connections.xml."""
    conn_xml = b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<connections xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <connection id="1" name="SQLServer_Finance"
              type="1" refreshedVersion="6" background="1"
              saveData="1">
    <dbPr connection="Provider=SQLOLEDB;Data Source=fin-db-01;Initial Catalog=FinanceDB;Integrated Security=SSPI;"
          command="SELECT * FROM dbo.MarketData WHERE Year=2024"
          commandType="2"/>
  </connection>
</connections>"""

    _patch_zip(xlsx_path, {"xl/connections.xml": conn_xml})


def _inject_external_links(
    xlsx_path: Path,
    target_file: str,
    formulas: list[tuple[str, str, str]],
) -> None:
    """Inject external link rels and patch formulas into sheet XML.

    formulas: list of (sheet_name, cell_ref, formula_text)
    """
    # Build the rels file for externalLink1
    ext_link_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<externalLink xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <externalBook xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:id="rId1">
    <sheetNames><sheetName val="Data"/><sheetName val="Lookup"/></sheetNames>
  </externalBook>
</externalLink>""".encode()

    ext_rels_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1"
    Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLinkPath"
    Target="{target_file}" TargetMode="External"/>
</Relationships>""".encode()

    patches = {
        "xl/externalLinks/externalLink1.xml": ext_link_xml,
        "xl/externalLinks/_rels/externalLink1.xml.rels": ext_rels_xml,
    }

    # Patch sheet XMLs to inject the formulas
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        # Build sheet name → zip path map
        import lxml.etree as ET
        wb_xml = zf.read("xl/workbook.xml")
        wb_root = ET.fromstring(wb_xml)
        ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

        sheet_map: dict[str, str] = {}
        rels_xml = zf.read("xl/_rels/workbook.xml.rels")
        rels: dict[str, str] = {}
        for rel in ET.fromstring(rels_xml):
            rid = rel.get("Id", "")
            target = rel.get("Target", "")
            if rid and target:
                rels[rid] = target

        for sh in (wb_root.findall(f".//{{{ns}}}sheet")
                   or wb_root.findall(".//{*}sheet")):
            name = sh.get("name", "")
            rid = sh.get(f"{{{rel_ns}}}id") or sh.get("r:id", "")
            if rid in rels:
                t = rels[rid].lstrip("/")
                if not t.startswith("xl/"):
                    t = f"xl/{t}"
                sheet_map[name] = t

        # For each formula, patch the sheet XML
        for sheet_name, cell_ref, formula_text in formulas:
            sheet_path = sheet_map.get(sheet_name)
            if not sheet_path or sheet_path not in zf.namelist():
                continue
            data = zf.read(sheet_path)
            # Find or create the cell element
            col_match = re.match(r"([A-Z]+)(\d+)", cell_ref)
            if not col_match:
                continue
            # Insert a new cell with the formula before </sheetData>
            cell_xml = (
                f'<c r="{cell_ref}">'
                f"<f>{formula_text}</f>"
                f"<v>0</v></c>"
            ).encode()
            # Insert before </sheetData>
            data = data.replace(
                b"</sheetData>",
                b"<row r=\"999\"><c r=\"" + cell_ref.encode() + b"\">"
                b"<f>" + formula_text.encode() + b"</f>"
                b"<v>0</v></c></row></sheetData>",
            )
            patches[sheet_path] = data

    _patch_zip(xlsx_path, patches)


def _patch_zip(xlsx_path: Path, patches: dict[str, bytes]) -> None:
    """Patch files inside a ZIP archive."""
    tmp = xlsx_path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(xlsx_path, "r") as zin, zipfile.ZipFile(tmp, "w") as zout:
        for item in zin.namelist():
            if item in patches:
                zout.writestr(item, patches[item])
            else:
                zout.writestr(item, zin.read(item))
        # Add new files that don't exist yet
        existing = set(zin.namelist())
        for name, data in patches.items():
            if name not in existing:
                zout.writestr(name, data)
    tmp.replace(xlsx_path)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def generate_all(out_dir: Path) -> tuple[Path, Path, Path]:
    """Generate all three fixture files.

    Returns (model_path, upstream_a_path, upstream_b_path).
    """
    out_dir.mkdir(parents=True, exist_ok=True)

    upstream_b = out_dir / "upstream_b.xlsx"
    upstream_a = out_dir / "upstream_a.xlsx"
    model = out_dir / "model.xlsx"

    _generate_upstream_b(upstream_b)
    _generate_upstream_a(upstream_a, upstream_b.name)
    _generate_model(model, upstream_a.name)

    return model, upstream_a, upstream_b


if __name__ == "__main__":
    import sys
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("BusinessContract/fixtures/data")
    m, a, b = generate_all(out)
    print(f"Generated:\n  {m}\n  {a}\n  {b}")
