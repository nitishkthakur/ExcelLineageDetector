"""Scan model file: extract vectors, formulas, connections, and external refs."""
from __future__ import annotations

import io
import logging
import re
import sys
import zipfile
from collections import defaultdict
from pathlib import Path

import openpyxl

# Reuse existing modules from the main project
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from lineage.detector import ExcelLineageDetector
from lineage.hardcoded_scanner import scan_vectors

from .config import ContractConfig
from .models import (
    BusinessContract,
    ContractVariable,
    DependencyEdge,
    TransformationStep,
)

logger = logging.getLogger(__name__)

_COL_RE = re.compile(r"^\$?([A-Z]+)\$?(\d+)$")


def _col_to_num(col: str) -> int:
    """Convert column letters to number: A->1, Z->26, AA->27."""
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _parse_cell(ref: str) -> tuple[str, int] | None:
    """Parse a cell ref like 'B3' or '$B$3' -> ('B', 3)."""
    m = _COL_RE.match(ref)
    if not m:
        return None
    return m.group(1), int(m.group(2))


def _parse_range(rng: str) -> tuple[int, int, int, int] | None:
    """Parse 'B2:M13' -> (min_col, min_row, max_col, max_row) as ints.

    Returns None if not a valid range or single cell.
    """
    rng = rng.replace("$", "")
    parts = rng.split(":")
    p1 = _parse_cell(parts[0])
    if not p1:
        return None
    if len(parts) == 1:
        col_num = _col_to_num(p1[0])
        return (col_num, p1[1], col_num, p1[1])
    p2 = _parse_cell(parts[1])
    if not p2:
        return None
    c1, c2 = _col_to_num(p1[0]), _col_to_num(p2[0])
    return (min(c1, c2), min(p1[1], p2[1]), max(c1, c2), max(p1[1], p2[1]))


def _ranges_overlap(ref: str, var_range: str) -> bool:
    """Check if ref (cell or range) overlaps with var_range."""
    r1 = _parse_range(ref)
    r2 = _parse_range(var_range)
    if not r1 or not r2:
        return False
    # Check rectangle overlap
    return not (r1[2] < r2[0] or r1[0] > r2[2] or r1[3] < r2[1] or r1[1] > r2[3])


# ── Connections ──────────────────────────────────────────────────


def _scan_connections(model_path: Path) -> list[dict]:
    """Run full detector to extract all DataConnections."""
    try:
        detector = ExcelLineageDetector()
        results = detector.detect(str(model_path))
    except Exception as e:
        logger.warning("Connection scan failed: %s", e)
        return []
    return [
        {
            "id": dc.id,
            "category": dc.category,
            "sub_type": dc.sub_type,
            "raw_connection": dc.raw_connection,
            "location": dc.location,
            "confidence": dc.confidence,
            "parsed_query": getattr(dc, "parsed_query", None),
        }
        for dc in results
    ]


# ── Formula scanning ─────────────────────────────────────────────


def _scan_formulas(model_path: Path, sheet_names: list[str]) -> list[dict]:
    """Extract all formulas from specified sheets using streaming XML."""
    from lxml import etree

    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    formulas = []

    with zipfile.ZipFile(str(model_path), "r") as zf:
        try:
            wb_xml = zf.read("xl/workbook.xml")
        except KeyError:
            return []

        sheet_rids: dict[str, str] = {}
        for _, elem in etree.iterparse(io.BytesIO(wb_xml), events=("end",)):
            if etree.QName(elem.tag).localname == "sheet":
                name = elem.get("name", "")
                rid = elem.get(
                    r"{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
                )
                if name in sheet_names and rid:
                    sheet_rids[rid] = name
                elem.clear()

        try:
            rels_xml = zf.read("xl/_rels/workbook.xml.rels")
        except KeyError:
            return []

        rid_to_path: dict[str, str] = {}
        for _, elem in etree.iterparse(io.BytesIO(rels_xml), events=("end",)):
            if etree.QName(elem.tag).localname == "Relationship":
                rid = elem.get("Id", "")
                target = elem.get("Target", "")
                if rid in sheet_rids:
                    path = "xl/" + target if not target.startswith("/") else target.lstrip("/")
                    rid_to_path[path] = sheet_rids[rid]
                elem.clear()

        for xml_path, sheet_name in rid_to_path.items():
            try:
                data = zf.read(xml_path)
            except KeyError:
                continue

            for _, elem in etree.iterparse(
                io.BytesIO(data), events=("end",), tag=f"{{{ns}}}c"
            ):
                ref = elem.get("r", "")
                f_elem = elem.find(f"{{{ns}}}f")
                if f_elem is not None and f_elem.text:
                    v_elem = elem.find(f"{{{ns}}}v")
                    formulas.append({
                        "sheet": sheet_name,
                        "cell": ref,
                        "formula": f_elem.text,
                        "value": v_elem.text if v_elem is not None else None,
                    })
                elem.clear()

    return formulas


# ── Named ranges / scalar scanning ───────────────────────────────


def _scan_named_ranges(model_path: Path) -> list[ContractVariable]:
    """Scan workbook defined names and emit scalar/vector variables."""
    from lxml import etree

    variables = []
    with zipfile.ZipFile(str(model_path), "r") as zf:
        try:
            wb_xml = zf.read("xl/workbook.xml")
        except KeyError:
            return []

        # Map localSheetId -> sheet name
        sheet_id_map: dict[str, str] = {}
        for _, elem in etree.iterparse(io.BytesIO(wb_xml), events=("end",)):
            local = etree.QName(elem.tag).localname
            if local == "sheet":
                sid = elem.get("sheetId", "")
                name = elem.get("name", "")
                # localSheetId in definedName is 0-based index, sheetId is 1-based
                # We need index-based mapping
                elem.clear()

        # Re-parse to get sheet names in order (index = position)
        wb_xml2 = zf.read("xl/workbook.xml")
        sheet_names_ordered: list[str] = []
        for _, elem in etree.iterparse(io.BytesIO(wb_xml2), events=("end",)):
            local = etree.QName(elem.tag).localname
            if local == "sheet":
                sheet_names_ordered.append(elem.get("name", ""))
                elem.clear()

        # Now parse defined names
        wb_xml3 = zf.read("xl/workbook.xml")
        ref_re = re.compile(r"^'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+)(?::\$?([A-Z]+)\$?(\d+))?$")

        for _, elem in etree.iterparse(io.BytesIO(wb_xml3), events=("end",)):
            local = etree.QName(elem.tag).localname
            if local != "definedName":
                elem.clear()
                continue

            name = elem.get("name", "")
            value = (elem.text or "").strip()
            hidden = elem.get("hidden", "0") == "1"
            elem.clear()

            # Skip hidden and built-in names
            if hidden or name.startswith("_xlnm"):
                continue

            # Skip formula-based (dynamic) names — contain functions
            if any(c in value for c in ("(", "OFFSET", "INDIRECT")):
                continue

            m = ref_re.match(value)
            if not m:
                continue

            sheet = m.group(1)
            start_col, start_row = m.group(2), int(m.group(3))
            end_col = m.group(4) or start_col
            end_row = int(m.group(5)) if m.group(5) else start_row

            cell_range = f"{start_col}{start_row}:{end_col}{end_row}"
            length = max(end_row - start_row + 1, _col_to_num(end_col) - _col_to_num(start_col) + 1)

            if length == 1:
                direction = "scalar"
            elif end_row - start_row >= _col_to_num(end_col) - _col_to_num(start_col):
                direction = "column"
            else:
                direction = "row"

            # Read sample values
            sample_vals: list[float] = []
            try:
                wb = openpyxl.load_workbook(str(model_path), data_only=True)
                ws = wb[sheet]
                if direction == "column" or direction == "scalar":
                    col_idx = _col_to_num(start_col)
                    for r in range(start_row, min(start_row + 5, end_row + 1)):
                        v = ws.cell(row=r, column=col_idx).value
                        if isinstance(v, (int, float)):
                            sample_vals.append(float(v))
                else:
                    for c in range(_col_to_num(start_col), min(_col_to_num(start_col) + 5, _col_to_num(end_col) + 1)):
                        v = ws.cell(row=start_row, column=c).value
                        if isinstance(v, (int, float)):
                            sample_vals.append(float(v))
                wb.close()
            except Exception:
                pass

            var = ContractVariable.from_vector(
                sheet=sheet,
                cell_range=cell_range,
                direction=direction,
                length=length,
                values=sample_vals,
                source_type="hardcoded",
            )
            var.business_name = name  # Named ranges already have business names!
            variables.append(var)

    return variables


# ── Hardcoded vectors ─────────────────────────────────────────────


def _scan_hardcoded_vectors(
    model_path: Path, sheet_names: list[str], min_length: int = 3
) -> list[ContractVariable]:
    """Scan for hardcoded numeric vectors using the existing scanner."""
    all_vectors = scan_vectors(model_path)
    variables = []
    for sheet_name in sheet_names:
        for vec in all_vectors.get(sheet_name, []):
            if vec.length >= min_length:
                var = ContractVariable.from_vector(
                    sheet=sheet_name,
                    cell_range=vec.cell_range,
                    direction=vec.direction,
                    length=vec.length,
                    values=vec.sample_values,
                    source_type="hardcoded",
                )
                variables.append(var)
    return variables


# ── Formula variables ─────────────────────────────────────────────


def _scan_formula_variables(
    formulas: list[dict], min_length: int = 3
) -> list[ContractVariable]:
    """Group formula cells into contiguous vector variables.

    Also emits scalar formula cells that don't belong to any vector.
    """
    # Group by (sheet, column_letter) for column vectors
    col_groups: dict[tuple[str, str], list[dict]] = defaultdict(list)
    # Group by (sheet, row_number) for row vectors
    row_groups: dict[tuple[str, int], list[dict]] = defaultdict(list)

    all_cells: dict[tuple[str, str], dict] = {}  # (sheet, cell_ref) -> formula dict

    for f in formulas:
        parsed = _parse_cell(f["cell"])
        if not parsed:
            continue
        col_letter, row_num = parsed
        col_groups[(f["sheet"], col_letter)].append(f)
        row_groups[(f["sheet"], row_num)].append(f)
        all_cells[(f["sheet"], f["cell"])] = f

    variables = []
    seen_ranges: set[str] = set()
    cells_in_vectors: set[tuple[str, str]] = set()  # track which cells are in vectors

    # Column vectors
    for (sheet, col_letter), cells in col_groups.items():
        if len(cells) < min_length:
            continue
        cells.sort(key=lambda c: _parse_cell(c["cell"])[1])
        rows = [_parse_cell(c["cell"])[1] for c in cells]
        if rows[-1] - rows[0] + 1 == len(rows):
            range_str = f"{col_letter}{rows[0]}:{col_letter}{rows[-1]}"
            key = f"{sheet}!{range_str}"
            if key not in seen_ranges:
                seen_ranges.add(key)
                vals = []
                for c in cells:
                    try:
                        vals.append(float(c["value"]))
                    except (TypeError, ValueError):
                        vals.append(0.0)
                var = ContractVariable.from_vector(
                    sheet=sheet,
                    cell_range=range_str,
                    direction="column",
                    length=len(cells),
                    values=vals,
                    source_type="formula",
                )
                var.excel_formula = cells[0]["formula"]
                variables.append(var)
                for c in cells:
                    cells_in_vectors.add((sheet, c["cell"]))

    # Row vectors
    for (sheet, row_num), cells in row_groups.items():
        if len(cells) < min_length:
            continue
        cells.sort(key=lambda c: _col_to_num(_parse_cell(c["cell"])[0]))
        col_indices = [_col_to_num(_parse_cell(c["cell"])[0]) for c in cells]
        if col_indices[-1] - col_indices[0] + 1 == len(col_indices):
            start_col = _parse_cell(cells[0]["cell"])[0]
            end_col = _parse_cell(cells[-1]["cell"])[0]
            range_str = f"{start_col}{row_num}:{end_col}{row_num}"
            key = f"{sheet}!{range_str}"
            if key not in seen_ranges:
                seen_ranges.add(key)
                vals = []
                for c in cells:
                    try:
                        vals.append(float(c["value"]))
                    except (TypeError, ValueError):
                        vals.append(0.0)
                var = ContractVariable.from_vector(
                    sheet=sheet,
                    cell_range=range_str,
                    direction="row",
                    length=len(cells),
                    values=vals,
                    source_type="formula",
                )
                var.excel_formula = cells[0]["formula"]
                variables.append(var)
                for c in cells:
                    cells_in_vectors.add((sheet, c["cell"]))

    # Scalar formula cells — cells not part of any vector
    for (sheet, cell_ref), f in all_cells.items():
        if (sheet, cell_ref) in cells_in_vectors:
            continue
        key = f"{sheet}!{cell_ref}"
        if key in seen_ranges:
            continue
        seen_ranges.add(key)
        try:
            val = float(f["value"]) if f["value"] else 0.0
        except (TypeError, ValueError):
            val = 0.0
        var = ContractVariable.from_vector(
            sheet=sheet,
            cell_range=cell_ref,
            direction="scalar",
            length=1,
            values=[val],
            source_type="formula",
        )
        var.excel_formula = f["formula"]
        variables.append(var)

    return variables


# ── Orchestrator ──────────────────────────────────────────────────


def scan_model(config: ContractConfig) -> BusinessContract:
    """Full model scan: vectors, formulas, connections, external refs."""
    model_path = config.model_path
    wb = openpyxl.load_workbook(str(model_path), data_only=True)
    all_sheets = wb.sheetnames
    wb.close()

    # 1. Connections
    connections = _scan_connections(model_path)

    # 2. Formulas from all sheets
    formulas = _scan_formulas(model_path, all_sheets)

    # 3. Hardcoded vectors
    hardcoded_vars = _scan_hardcoded_vectors(
        model_path, all_sheets, config.min_vector_length
    )

    # 4. Formula vectors + scalars
    formula_vars = _scan_formula_variables(formulas, config.min_vector_length)

    # 5. Named ranges (scalars and vectors from defined names)
    named_vars = _scan_named_ranges(model_path)

    # 6. Deduplicate: named ranges may overlap with hardcoded/formula vars
    all_vars = hardcoded_vars + formula_vars
    existing_locations = {v.excel_location for v in all_vars}
    for nv in named_vars:
        if nv.excel_location not in existing_locations:
            all_vars.append(nv)
            existing_locations.add(nv.excel_location)
        else:
            # Merge the business_name from the named range onto the existing var
            for v in all_vars:
                if v.excel_location == nv.excel_location and not v.business_name:
                    v.business_name = nv.business_name

    # 7. Mark output variables
    for var in all_vars:
        if var.sheet in config.output_sheets:
            var.variable_type = "output"
        elif var.source_type == "hardcoded":
            var.variable_type = "input"

    # 8. Build transformation steps
    transformations = _build_transformations(formulas, all_vars)

    # 9. Build dependency edges (deduplicated)
    edges = _build_edges(transformations)

    contract = BusinessContract(
        model_file=str(model_path),
        output_sheets=config.output_sheets,
        variables=all_vars,
        transformations=transformations,
        edges=edges,
        connections=connections,
    )

    return contract


# ── Transformation building ───────────────────────────────────────


def _build_transformations(
    formulas: list[dict], variables: list[ContractVariable]
) -> list[TransformationStep]:
    """Build transformation steps linking formula variables to their inputs.

    Uses indexed lookups (O(1) per ref) instead of scanning all variables.
    """
    # Index: (sheet, cell_ref) -> variable id (for start cells of ranges)
    cell_to_var: dict[tuple[str, str], str] = {}
    for var in variables:
        start = var.cell_range.split(":")[0].replace("$", "")
        cell_to_var[(var.sheet, start)] = var.id

    # Index: (sheet) -> list of (parsed_range, var_id) for overlap checks
    sheet_ranges: dict[str, list[tuple[tuple[int, int, int, int], str]]] = defaultdict(list)
    for var in variables:
        parsed = _parse_range(var.cell_range)
        if parsed:
            sheet_ranges[var.sheet].append((parsed, var.id))

    # Regex for cell/range references in formulas (handles Sheet!Ref and bare Ref)
    ref_re = re.compile(r"(?:'?([^'!(),\[\]]+)'?!)?(\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?)")

    steps = []
    seen_outputs: set[str] = set()

    for f in formulas:
        output_key = (f["sheet"], f["cell"].replace("$", ""))
        output_var_id = cell_to_var.get(output_key)
        if not output_var_id or output_var_id in seen_outputs:
            continue
        seen_outputs.add(output_var_id)

        # Find input variable references in formula
        input_ids: list[str] = []
        input_id_set: set[str] = set()

        for match in ref_re.finditer(f["formula"]):
            ref_sheet = match.group(1) or f["sheet"]
            ref_str = match.group(2).replace("$", "")
            ref_parsed = _parse_range(ref_str)
            if not ref_parsed:
                continue

            # O(n) within sheet but typically small per-sheet list
            for var_range, var_id in sheet_ranges.get(ref_sheet, []):
                if var_id in input_id_set or var_id == output_var_id:
                    continue
                # Check overlap between ref and variable range
                if not (ref_parsed[2] < var_range[0] or ref_parsed[0] > var_range[2] or
                        ref_parsed[3] < var_range[1] or ref_parsed[1] > var_range[3]):
                    input_ids.append(var_id)
                    input_id_set.add(var_id)

        step = TransformationStep.make(
            output_id=output_var_id,
            input_ids=input_ids,
            excel_formula=f["formula"],
            sql_formula="",
            sheet=f["sheet"],
            cell_range=f["cell"],
        )
        steps.append(step)

    return steps


def _build_edges(
    transformations: list[TransformationStep],
) -> list[DependencyEdge]:
    """Build deduplicated dependency edges from transformations."""
    seen: set[tuple[str, str]] = set()
    edges = []

    for step in transformations:
        for input_id in step.input_variable_ids:
            key = (input_id, step.output_variable_id)
            if key not in seen:
                seen.add(key)
                edges.append(DependencyEdge(
                    source_id=input_id,
                    target_id=step.output_variable_id,
                    edge_type="formula",
                    metadata={"formula": step.excel_formula},
                ))

    return edges
