"""Write the Business Contract to an Excel file."""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import BusinessContract, ContractVariable, DependencyEdge, TransformationStep

_HEADER_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
_INPUT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_OUTPUT_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
_FORMULA_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


def write_contract(contract: BusinessContract, out_path: Path) -> Path:
    """Write the complete Business Contract Excel file."""
    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary(ws_summary, contract)

    # Sheet 2: Variables
    ws_vars = wb.create_sheet("Variables")
    _write_variables(ws_vars, contract.variables)

    # Sheet 3: Transformations
    ws_tx = wb.create_sheet("Transformations")
    _write_transformations(ws_tx, contract.transformations, contract.variables)

    # Sheet 4: Dependencies
    ws_deps = wb.create_sheet("Dependencies")
    _write_dependencies(ws_deps, contract.edges, contract.variables)

    # Sheet 5: Connections
    if contract.connections:
        ws_conn = wb.create_sheet("External Connections")
        _write_connections(ws_conn, contract.connections)

    # Sheet 6: Upstream Lineage
    upstream_vars = [v for v in contract.variables if v.upstream_source]
    if upstream_vars:
        ws_up = wb.create_sheet("Upstream Lineage")
        _write_upstream(ws_up, upstream_vars)

    wb.save(str(out_path))
    return out_path


def _apply_header(ws, row: int, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = _HEADER_FONT_WHITE
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws, min_width: int = 10, max_width: int = 50) -> None:
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def _write_summary(ws, contract: BusinessContract) -> None:
    ws.cell(row=1, column=1, value="Business Contract Summary").font = Font(bold=True, size=14)
    ws.cell(row=3, column=1, value="Model File:").font = _HEADER_FONT
    ws.cell(row=3, column=2, value=contract.model_file)
    ws.cell(row=4, column=1, value="Output Sheets:").font = _HEADER_FONT
    ws.cell(row=4, column=2, value=", ".join(contract.output_sheets))
    ws.cell(row=5, column=1, value="Total Variables:").font = _HEADER_FONT
    ws.cell(row=5, column=2, value=len(contract.variables))

    inputs = sum(1 for v in contract.variables if v.variable_type == "input")
    outputs = sum(1 for v in contract.variables if v.variable_type == "output")
    intermediates = len(contract.variables) - inputs - outputs

    ws.cell(row=6, column=1, value="Input Variables:").font = _HEADER_FONT
    ws.cell(row=6, column=2, value=inputs)
    ws.cell(row=7, column=1, value="Output Variables:").font = _HEADER_FONT
    ws.cell(row=7, column=2, value=outputs)
    ws.cell(row=8, column=1, value="Intermediate Variables:").font = _HEADER_FONT
    ws.cell(row=8, column=2, value=intermediates)
    ws.cell(row=9, column=1, value="Transformations:").font = _HEADER_FONT
    ws.cell(row=9, column=2, value=len(contract.transformations))
    ws.cell(row=10, column=1, value="External Connections:").font = _HEADER_FONT
    ws.cell(row=10, column=2, value=len(contract.connections))

    _auto_width(ws)


def _write_variables(ws, variables: list[ContractVariable]) -> None:
    headers = [
        "ID", "Business Name", "Sheet", "Cell Range", "Direction",
        "Length", "Type", "Source Type", "Sample Values",
        "Upstream Source", "Confidence", "Match Type",
    ]
    _apply_header(ws, 1, headers)

    type_fills = {
        "input": _INPUT_FILL,
        "output": _OUTPUT_FILL,
        "intermediate": _FORMULA_FILL,
    }

    for row_idx, var in enumerate(variables, 2):
        data = [
            var.id, var.business_name, var.sheet, var.cell_range,
            var.direction, var.length, var.variable_type, var.source_type,
            str(var.sample_values[:5]),
            var.upstream_source or "",
            var.confidence if var.confidence else "",
            var.match_type or "",
        ]
        fill = type_fills.get(var.variable_type)
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if fill:
                cell.fill = fill

    _auto_width(ws)


def _write_transformations(
    ws, transformations: list[TransformationStep],
    variables: list[ContractVariable],
) -> None:
    var_map = {v.id: v.business_name or v.id for v in variables}

    headers = [
        "ID", "Output Variable", "Input Variables",
        "Excel Formula", "SQL Formula", "Sheet", "Cell Range",
    ]
    _apply_header(ws, 1, headers)

    for row_idx, tx in enumerate(transformations, 2):
        output_name = var_map.get(tx.output_variable_id, tx.output_variable_id)
        input_names = ", ".join(var_map.get(iid, iid) for iid in tx.input_variable_ids)

        data = [
            tx.id, output_name, input_names,
            tx.excel_formula, tx.sql_formula, tx.sheet, tx.cell_range,
        ]
        for col_idx, val in enumerate(data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    _auto_width(ws)


def _write_dependencies(
    ws, edges: list[DependencyEdge],
    variables: list[ContractVariable],
) -> None:
    var_map = {v.id: v.business_name or v.id for v in variables}

    headers = ["Source", "Target", "Edge Type", "Details"]
    _apply_header(ws, 1, headers)

    for row_idx, edge in enumerate(edges, 2):
        source_name = var_map.get(edge.source_id, edge.source_id)
        target_name = var_map.get(edge.target_id, edge.target_id)

        data = [
            source_name, target_name, edge.edge_type,
            str(edge.metadata) if edge.metadata else "",
        ]
        for col_idx, val in enumerate(data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    _auto_width(ws)


def _write_connections(ws, connections: list[dict]) -> None:
    headers = ["ID", "Category", "Sub-Type", "Raw Connection", "Location", "Confidence"]
    _apply_header(ws, 1, headers)

    for row_idx, conn in enumerate(connections, 2):
        data = [
            conn.get("id", ""),
            conn.get("category", ""),
            conn.get("sub_type", ""),
            conn.get("raw_connection", ""),
            conn.get("location", ""),
            conn.get("confidence", ""),
        ]
        for col_idx, val in enumerate(data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    _auto_width(ws)


def _write_upstream(ws, variables: list[ContractVariable]) -> None:
    headers = [
        "Business Name", "Sheet", "Cell Range",
        "Upstream File", "Upstream Sheet", "Upstream Range",
        "Confidence", "Match Type",
    ]
    _apply_header(ws, 1, headers)

    for row_idx, var in enumerate(variables, 2):
        data = [
            var.business_name, var.sheet, var.cell_range,
            var.upstream_file or "", var.upstream_sheet or "",
            var.upstream_range or "",
            var.confidence, var.match_type or "",
        ]
        for col_idx, val in enumerate(data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    _auto_width(ws)
