"""Excel reporter for Excel Lineage Detector."""

from __future__ import annotations

import re
from pathlib import Path

from lineage.models import DataConnection


# Category color coding (background fill colors).
# Add an entry here whenever a new category is introduced.
CATEGORY_COLORS = {
    "database":   "E3F2FD",
    "file":       "E8F5E9",
    "web":        "FFF3E0",
    "powerquery": "F3E5F5",
    "vba":        "FFEBEE",
    "pivot":      "E0F7FA",
    "formula":    "E8EAF6",
    "hyperlink":  "FFF8E1",
    "ole":        "F9FBE7",
    "metadata":   "FAFAFA",
    "input":      "FFF9C4",
}

HEADER_FILL_COLOR = "1565C0"
HEADER_TEXT_COLOR = "FFFFFF"

_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def _safe_sheet_name(name: str, used: set[str], max_len: int = 31) -> str:
    """Sanitize a sheet name and ensure uniqueness within *used*."""
    safe = _INVALID_SHEET_CHARS.sub("_", name).strip()[:max_len] or "Sheet"
    candidate = safe
    suffix = 1
    while candidate in used:
        sfx = f"_{suffix}"
        candidate = safe[: max_len - len(sfx)] + sfx
        suffix += 1
    used.add(candidate)
    return candidate


def _fmt_val(v: object) -> str:
    """Format a sample numeric value compactly."""
    if isinstance(v, float):
        if v == int(v) and abs(v) < 1e12:
            return str(int(v))
        return f"{v:.4g}"
    return str(v)


class ExcelReporter:
    """Creates a formatted Excel report of lineage findings.

    Output structure
    ----------------
    Sheet 1  – All Connections   (every DataConnection, colour-coded by category)
    Sheet 2+ – one sheet per sheet in the analysed workbook, listing all
               hardcoded-value vectors (contiguous runs of non-formula numerics)
               found on that sheet.
    """

    def write(
        self,
        connections: list[DataConnection],
        input_path: Path,
        out_dir: Path,
    ) -> Path:
        """Write the lineage report to *out_dir* and return its path."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise ImportError(f"openpyxl is required for Excel reporting: {e}")

        stem = input_path.stem
        out = out_dir / f"{stem}_lineage_report.xlsx"

        wb = Workbook()
        wb.remove(wb.active)

        # ── shared style helpers (closures) ──────────────────────────────────
        def hdr_fill():
            return PatternFill("solid", fgColor=HEADER_FILL_COLOR)

        def hdr_font():
            return Font(bold=True, color=HEADER_TEXT_COLOR, size=11)

        def cat_fill(category: str):
            return PatternFill("solid", fgColor=CATEGORY_COLORS.get(category, "FFFFFF"))

        def thin_border():
            s = Side(style="thin", color="CCCCCC")
            return Border(left=s, right=s, top=s, bottom=s)

        def hdr_border():
            s = Side(style="medium", color="FFFFFF")
            return Border(left=s, right=s, top=s, bottom=s)

        def style_header(ws, headers: list[str], row: int = 1):
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=ci, value=h)
                cell.fill = hdr_fill()
                cell.font = hdr_font()
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = hdr_border()
            ws.row_dimensions[row].height = 22

        def auto_width(ws, min_w: int = 10, max_w: int = 60):
            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                max_len = max(
                    (len(str(c.value)) for c in col if c.value),
                    default=0,
                )
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

        used_names: set[str] = set()

        # ── Sheet 1: All Connections ─────────────────────────────────────────
        ws_all = wb.create_sheet(_safe_sheet_name("All Connections", used_names))
        self._write_all_connections(ws_all, connections, style_header, cat_fill,
                                    thin_border, auto_width, get_column_letter)

        # ── Sheets 2-N: per-original-sheet hardcoded vector sheets ───────────
        try:
            from lineage.hardcoded_scanner import scan_vectors
            vectors_by_sheet = scan_vectors(input_path)
            for sheet_name, vectors in vectors_by_sheet.items():
                safe = _safe_sheet_name(sheet_name, used_names)
                ws = wb.create_sheet(safe)
                self._write_vector_sheet(ws, sheet_name, vectors,
                                         thin_border, get_column_letter)
        except Exception:
            pass

        wb.save(str(out))
        return out

    # -------------------------------------------------------------------------
    # All Connections sheet
    # -------------------------------------------------------------------------

    def _write_all_connections(self, ws, connections, style_header, cat_fill,
                                thin_border, auto_width, get_column_letter):
        from openpyxl.styles import Alignment

        headers = ["ID", "Category", "Sub-type", "Source", "Location",
                   "Query/Formula", "Confidence", "Raw Connection"]
        style_header(ws, headers)
        ws.freeze_panes = "A2"

        for i, conn in enumerate(connections, 2):
            fill = cat_fill(conn.category)
            row_data = [
                conn.id,
                conn.category,
                conn.sub_type,
                conn.source,
                conn.location,
                (conn.query_text or "")[:200] if conn.query_text else "",
                round(conn.confidence, 2),
                conn.raw_connection[:200],
            ]
            for ci, value in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=ci, value=value)
                cell.fill = fill
                cell.border = thin_border()
                if ci == 6:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        auto_width(ws)
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["D"].width = 35
        ws.column_dimensions["E"].width = 30
        ws.column_dimensions["F"].width = 45
        ws.column_dimensions["H"].width = 40

    # -------------------------------------------------------------------------
    # Per-sheet hardcoded vector sheet
    # -------------------------------------------------------------------------

    def _write_vector_sheet(self, ws, sheet_name: str, vectors,
                             thin_border, get_column_letter):
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # Title row
        ws.cell(row=1, column=1, value=f"Sheet: {sheet_name}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="1565C0")
        ws.row_dimensions[1].height = 20

        if not vectors:
            ws.cell(row=3, column=1,
                    value="No hardcoded value vectors found in this sheet.")
            ws.cell(row=3, column=1).font = Font(italic=True, color="888888")
            ws.column_dimensions["A"].width = 52
            return

        # Headers at row 3 (row 2 left blank as spacer)
        headers = ["Sheet Name", "Cell Range", "Direction", "Length",
                   "Start Cell", "End Cell", "Sample Values (first 5)"]
        hdr_fill_obj = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
        hdr_font_obj = Font(bold=True, color=HEADER_TEXT_COLOR, size=11)
        hdr_border_obj = Border(
            left=Side(style="medium", color="FFFFFF"),
            right=Side(style="medium", color="FFFFFF"),
        )
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.fill = hdr_fill_obj
            cell.font = hdr_font_obj
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = hdr_border_obj
        ws.row_dimensions[3].height = 22
        ws.freeze_panes = "A4"

        col_fill = PatternFill("solid", fgColor="E3F2FD")   # blue  → column vectors
        row_fill = PatternFill("solid", fgColor="E8F5E9")   # green → row vectors
        border = thin_border()

        for i, vec in enumerate(vectors, 4):
            fill = col_fill if vec.direction == "column" else row_fill

            sample_parts = [_fmt_val(v) for v in vec.sample_values]
            sample_str = ", ".join(sample_parts)
            if vec.length > len(vec.sample_values):
                sample_str += f", … ({vec.length - len(vec.sample_values)} more)"

            row_data = [
                vec.sheet,
                vec.cell_range,
                vec.direction,
                vec.length,
                vec.start_cell,
                vec.end_cell,
                sample_str,
            ]
            for ci, value in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=ci, value=value)
                cell.fill = fill
                cell.border = border
                if ci == 4:
                    cell.alignment = Alignment(horizontal="right")

        ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}3"

        # Fixed column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 48
