"""Excel report writer for upstream tracing results."""
from __future__ import annotations

import json
from pathlib import Path
from typing import TYPE_CHECKING

from lineage.tracing.config import TraceConfig
from lineage.tracing.models import TracingVector, VectorMatch

if TYPE_CHECKING:
    from lineage.tracing.formula_tracer import ExternalReference


# Colour palette
_HDR_FILL = "1565C0"
_HDR_TEXT = "FFFFFF"
_EXACT_FILL = "C8E6C9"           # green
_EXACT_SUB_FILL = "DCEDC8"       # light green
_APPROX_HIGH_FILL = "BBDEFB"     # blue (high similarity)
_APPROX_MED_FILL = "E3F2FD"      # lighter blue
_UNMATCHED_FILL = "FFF9C4"       # yellow
_ALT_TINT = "F5F5F5"             # alternating model-vector group
_FOUND_FILL = "C8E6C9"           # green — file found on disk
_MISSING_FILL = "FFCDD2"         # red/salmon — file not found


def _fmt_sample(values: list[float], total: int) -> str:
    """Format sample values compactly."""
    parts = []
    for v in values[:5]:
        if isinstance(v, float) and v == int(v) and abs(v) < 1e12:
            parts.append(str(int(v)))
        else:
            parts.append(f"{v:.4g}")
    s = ", ".join(parts)
    if total > 5:
        s += f", ... ({total - 5} more)"
    return s


class TracingReporter:
    """Writes upstream tracing results to a formatted Excel workbook."""

    def write(
        self,
        matches: list[VectorMatch],
        unmatched: list[TracingVector],
        config: TraceConfig,
        model_path: Path,
        sheet_name: str,
        upstream_files: list[Path],
        out_dir: Path,
    ) -> Path:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise ImportError(f"openpyxl is required for Excel reporting: {e}")

        stem = model_path.stem
        out = out_dir / f"upstream_tracing_{stem}.xlsx"

        wb = Workbook()
        wb.remove(wb.active)

        # Shared styles
        hdr_fill = PatternFill("solid", fgColor=_HDR_FILL)
        hdr_font = Font(bold=True, color=_HDR_TEXT, size=11)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_border = Border(
            left=Side(style="medium", color="FFFFFF"),
            right=Side(style="medium", color="FFFFFF"),
            top=Side(style="medium", color="FFFFFF"),
            bottom=Side(style="medium", color="FFFFFF"),
        )

        # ── Sheet 1: Config ──────────────────────────────────────────────
        ws_cfg = wb.create_sheet("Config")
        self._write_config_sheet(
            ws_cfg, config, model_path, sheet_name, upstream_files,
            matches, unmatched, hdr_fill, hdr_font, hdr_align, hdr_border, border,
        )

        # ── Sheet 2: Tracing Results ─────────────────────────────────────
        ws_res = wb.create_sheet("Tracing Results")
        self._write_results_sheet(
            ws_res, matches, unmatched,
            hdr_fill, hdr_font, hdr_align, hdr_border, border, get_column_letter,
        )

        wb.save(str(out))
        return out

    # ------------------------------------------------------------------ #
    # Config sheet
    # ------------------------------------------------------------------ #

    def _write_config_sheet(
        self, ws, config, model_path, sheet_name, upstream_files,
        matches, unmatched, hdr_fill, hdr_font, hdr_align, hdr_border, border,
    ):
        from openpyxl.styles import Font as F, Alignment as A

        ws.cell(row=1, column=1, value="Upstream Tracing Configuration")
        ws.cell(row=1, column=1).font = F(bold=True, size=14, color="1565C0")
        ws.row_dimensions[1].height = 24

        rows = [
            ("Model file", str(model_path)),
            ("Traced sheet", sheet_name),
            ("Upstream files", ", ".join(p.name for p in upstream_files)),
            ("", ""),
            ("Exact matching", "Enabled" if config.exact else "Disabled"),
            ("Approximate matching", "Enabled" if config.approximate else "Disabled"),
            ("Top-N (approximate)", str(config.top_n)),
            ("Similarity metric", config.similarity_metric),
            ("Min similarity", str(config.min_similarity)),
            ("Subsequence matching", "Yes" if config.subsequence_matching else "No"),
            ("Decimal places (exact)", str(config.exact_decimal_places)),
            ("Length tolerance %", f"{config.length_tolerance_pct}%"),
            ("Direction sensitive", "Yes" if config.direction_sensitive else "No"),
            ("Min vector length", str(config.min_vector_length)),
            ("", ""),
            ("Model vectors found", str(len(set(
                m.model_range for m in matches
            )) + len(unmatched))),
            ("Vectors with matches", str(len(set(m.model_range for m in matches)))),
            ("Vectors unmatched", str(len(unmatched))),
            ("Total match rows", str(len(matches))),
        ]

        for i, (label, value) in enumerate(rows, 3):
            c1 = ws.cell(row=i, column=1, value=label)
            c2 = ws.cell(row=i, column=2, value=value)
            c1.font = F(bold=True)
            c1.border = border
            c2.border = border

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 80

    # ------------------------------------------------------------------ #
    # Results sheet
    # ------------------------------------------------------------------ #

    def _write_results_sheet(
        self, ws, matches, unmatched,
        hdr_fill, hdr_font, hdr_align, hdr_border, border, get_column_letter,
    ):
        from openpyxl.styles import PatternFill, Font, Alignment

        headers = [
            "Model Range",           # A
            "Model Direction",       # B
            "Model Length",          # C
            "Model Sample Values",   # D
            "Match Rank",            # E
            "Match Type",            # F
            "Similarity",            # G
            "Upstream File",         # H
            "Upstream Sheet",        # I
            "Upstream Range",        # J
            "Upstream Matched Range",# K
            "Upstream Direction",    # L
            "Upstream Length",       # M
            "Upstream Sample Values",# N
        ]

        # Header row
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align
            cell.border = hdr_border
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        row_num = 2
        prev_model_range = None
        alt = False  # alternating tint for model vector groups

        # Write matched vectors
        for m in matches:
            if m.model_range != prev_model_range:
                alt = not alt
                prev_model_range = m.model_range

            fill = self._pick_fill(m.match_type, m.similarity, alt)

            row_data = [
                m.model_range,
                m.model_direction,
                m.model_length,
                _fmt_sample(m.model_sample, m.model_length),
                m.match_rank,
                m.match_type,
                m.similarity,
                m.upstream_file,
                m.upstream_sheet,
                m.upstream_range,
                m.upstream_matched_range,
                m.upstream_direction,
                m.upstream_length,
                _fmt_sample(m.upstream_sample, m.upstream_length),
            ]
            for ci, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=ci, value=value)
                cell.fill = fill
                cell.border = border
                if ci == 7:  # Similarity
                    cell.number_format = "0.000000"
                    cell.alignment = Alignment(horizontal="right")
                elif ci in (3, 5, 13):  # numeric columns
                    cell.alignment = Alignment(horizontal="right")
            row_num += 1

        # Write unmatched vectors
        unmatched_fill = PatternFill("solid", fgColor=_UNMATCHED_FILL)
        for u in unmatched:
            alt = not alt
            row_data = [
                u.cell_range,
                u.direction,
                u.length,
                _fmt_sample(list(u.values[:5]), u.length),
                "",             # rank
                "no match",     # type
                "",             # similarity
                "", "", "", "", "", "", "",
            ]
            for ci, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=ci, value=value)
                cell.fill = unmatched_fill
                cell.border = border
            row_num += 1

        # Auto-filter and column widths
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        widths = {
            "A": 16, "B": 12, "C": 10, "D": 38, "E": 10, "F": 18, "G": 12,
            "H": 28, "I": 20, "J": 16, "K": 20, "L": 12, "M": 10, "N": 38,
        }
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

    # ------------------------------------------------------------------ #
    # Formula tracing: Level N sheets
    # ------------------------------------------------------------------ #

    def write_with_levels(
        self,
        matches: list[VectorMatch],
        unmatched: list[TracingVector],
        config: TraceConfig,
        model_path: Path,
        sheet_name: str,
        upstream_files: list[Path],
        out_dir: Path,
        level_refs: dict[int, list[ExternalReference]] | None = None,
    ) -> Path:
        """Write tracing results AND formula-level sheets to one workbook.

        If *level_refs* is provided, adds one sheet per level ("Level 1",
        "Level 2", ...) after the value-tracing sheets.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise ImportError(f"openpyxl is required for Excel reporting: {e}")

        stem = model_path.stem
        out = out_dir / f"upstream_tracing_{stem}.xlsx"

        wb = Workbook()
        wb.remove(wb.active)

        # Shared styles
        hdr_fill = PatternFill("solid", fgColor=_HDR_FILL)
        hdr_font = Font(bold=True, color=_HDR_TEXT, size=11)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_border = Border(
            left=Side(style="medium", color="FFFFFF"),
            right=Side(style="medium", color="FFFFFF"),
            top=Side(style="medium", color="FFFFFF"),
            bottom=Side(style="medium", color="FFFFFF"),
        )

        # ── Config sheet ───────────────────────────────────────────────
        ws_cfg = wb.create_sheet("Config")
        self._write_config_sheet(
            ws_cfg, config, model_path, sheet_name, upstream_files,
            matches, unmatched, hdr_fill, hdr_font, hdr_align, hdr_border, border,
        )

        # ── Upstream Sources summary sheet ─────────────────────────────
        ws_sources = wb.create_sheet("Upstream Sources")
        self._write_sources_sheet(
            ws_sources, matches, level_refs,
            model_path,
            hdr_fill, hdr_font, hdr_align, hdr_border, border, get_column_letter,
        )

        # ── Tracing Results sheet (only if we have value-tracing data) ─
        if matches or unmatched:
            ws_res = wb.create_sheet("Tracing Results")
            self._write_results_sheet(
                ws_res, matches, unmatched,
                hdr_fill, hdr_font, hdr_align, hdr_border, border, get_column_letter,
            )

        # ── Level N sheets ─────────────────────────────────────────────
        if level_refs:
            for level in sorted(level_refs.keys()):
                ws = wb.create_sheet(f"Level {level}")
                self._write_level_sheet(
                    ws, level, level_refs[level],
                    hdr_fill, hdr_font, hdr_align, hdr_border, border,
                    get_column_letter,
                )

        wb.save(str(out))
        return out

    @staticmethod
    def _fmt_chain(chain: list[tuple[str, str, str]] | None) -> str:
        """Format a precedent chain for display."""
        if not chain:
            return "(direct)"
        parts = []
        for sheet, cell, formula in chain:
            snippet = formula[:80]
            if len(formula) > 80:
                snippet += "..."
            parts.append(f"{sheet}!{cell} (={snippet})")
        return " → ".join(parts)

    def _write_level_sheet(
        self, ws, level: int, refs: list[ExternalReference],
        hdr_fill, hdr_font, hdr_align, hdr_border, border, get_column_letter,
    ):
        """Write one "Level N" sheet with formula external references."""
        from openpyxl.styles import PatternFill, Font, Alignment

        headers = [
            "Source File",       # A
            "Source Sheet",      # B
            "Source Cell",       # C
            "Formula",           # D
            "Target File",       # E
            "Target Sheet",      # F
            "Target Range",      # G
            "Target Path",       # H
            "File Found",        # I
            "Resolved Path",     # J
            "Precedent Chain",   # K
            "Ref Type",          # L
            "Target Name",       # M
        ]

        # Header row
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align
            cell.border = hdr_border
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        found_fill = PatternFill("solid", fgColor=_FOUND_FILL)
        missing_fill = PatternFill("solid", fgColor=_MISSING_FILL)

        for ri, ref in enumerate(refs, 2):
            fill = found_fill if ref.file_found else missing_fill
            row_data = [
                ref.source_file,
                ref.source_sheet,
                ref.source_cell,
                ref.formula,
                ref.target_file,
                ref.target_sheet,
                ref.target_range,
                ref.target_path,
                "Yes" if ref.file_found else "No",
                ref.resolved_path,
                self._fmt_chain(ref.precedent_chain),
                ref.ref_type,
                ref.target_name or "",
            ]
            for ci, value in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=value)
                cell.fill = fill
                cell.border = border

        # Column widths
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        widths = {
            "A": 24, "B": 20, "C": 10, "D": 60, "E": 24,
            "F": 20, "G": 14, "H": 50, "I": 10, "J": 50, "K": 60,
            "L": 14, "M": 24,
        }
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

    # ------------------------------------------------------------------ #
    # Upstream Sources summary sheet
    # ------------------------------------------------------------------ #

    def _write_sources_sheet(
        self, ws, matches, level_refs, model_path,
        hdr_fill, hdr_font, hdr_align, hdr_border, border, get_column_letter,
    ):
        """Write a consolidated 'Upstream Sources' sheet.

        Gathers all unique upstream sources from:
        - Value-based matching (exact/approximate)
        - Formula tracing (all levels — already scoped to the selected sheet)

        Each row is a unique source with a 'File on Disk' status column.
        Only includes ancestors that feed the user-selected model sheet.
        """
        from openpyxl.styles import PatternFill, Font, Alignment

        headers = [
            "Source Name",           # A — filename, connection name, or URL
            "Source Type",           # B — value_match | formula_level_N | oledb | odbc | powerquery | ...
            "Category",             # C — file | database | web | powerquery | ...
            "Details",              # D — sheet, range, connection string snippet, etc.
            "Location in Model",    # E — where in the model this reference was found
            "File on Disk",         # F — Yes / No / N/A
        ]

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align
            cell.border = hdr_border
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        found_fill = PatternFill("solid", fgColor=_FOUND_FILL)
        missing_fill = PatternFill("solid", fgColor=_MISSING_FILL)
        na_fill = PatternFill("solid", fgColor=_ALT_TINT)

        # Collect rows: list of (name, source_type, category, details, location, file_on_disk)
        # file_on_disk: True, False, or None (N/A for databases/web)
        rows: list[tuple[str, str, str, str, str, bool | None]] = []
        seen_keys: set[str] = set()  # dedup key

        # ── 1. Value-based matches ────────────────────────────────────
        if matches:
            # Group by upstream file to show one row per file with best match type
            file_info: dict[str, tuple[str, list[str], list[str]]] = {}
            for m in matches:
                if not m.match_type.startswith("exact") and m.match_type != "approximate":
                    continue
                fn = m.upstream_file
                if fn not in file_info:
                    file_info[fn] = (m.match_type, [], [])
                _, sheets, ranges = file_info[fn]
                if m.upstream_sheet not in sheets:
                    sheets.append(m.upstream_sheet)
                if m.model_range not in ranges:
                    ranges.append(m.model_range)
                # Prefer exact over approximate
                if m.match_type.startswith("exact") and not file_info[fn][0].startswith("exact"):
                    file_info[fn] = (m.match_type, sheets, ranges)

            for fn, (match_type, sheets, model_ranges) in file_info.items():
                key = f"value|{fn}"
                if key in seen_keys:
                    continue
                seen_keys.add(key)
                file_found = self._check_file_exists(fn, model_path.parent)
                rows.append((
                    fn,
                    f"value_match ({match_type})",
                    "file",
                    f"Sheet(s): {', '.join(sheets)}",
                    f"Model range(s): {', '.join(model_ranges[:5])}",
                    file_found,
                ))

        # ── 2. Formula tracing refs (all levels) ─────────────────────
        if level_refs:
            for level in sorted(level_refs.keys()):
                # Group by target file for this level
                level_files: dict[str, tuple[bool, list[str], str]] = {}
                for ref in level_refs[level]:
                    fn = ref.target_file
                    if fn not in level_files:
                        level_files[fn] = (ref.file_found, [], ref.source_file)
                    _, sheets, _ = level_files[fn]
                    if ref.target_sheet not in sheets:
                        sheets.append(ref.target_sheet)

                for fn, (file_found, sheets, src_file) in level_files.items():
                    key = f"formula|{fn}"
                    if key in seen_keys:
                        continue
                    seen_keys.add(key)
                    rows.append((
                        fn,
                        f"formula_level_{level}",
                        "file",
                        f"Sheet(s): {', '.join(sheets)}",
                        f"Referenced from {src_file}",
                        file_found,
                    ))

        # ── Write rows ────────────────────────────────────────────────
        for ri, (name, src_type, category, details, location, on_disk) in enumerate(rows, 2):
            if on_disk is True:
                fill = found_fill
                disk_str = "Yes"
            elif on_disk is False:
                fill = missing_fill
                disk_str = "No"
            else:
                fill = na_fill
                disk_str = "N/A"

            row_data = [name, src_type, category, details, location, disk_str]
            for ci, value in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=value)
                cell.fill = fill
                cell.border = border

        # Column widths and auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        widths = {"A": 30, "B": 22, "C": 14, "D": 60, "E": 30, "F": 12}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

    @staticmethod
    def _check_file_exists(name_or_path: str, search_dir: Path) -> bool:
        """Check if a file exists on disk by name or path."""
        from pathlib import Path as P
        p = P(name_or_path)
        if p.is_absolute() and p.exists():
            return True
        candidate = search_dir / p.name
        if candidate.exists():
            return True
        # Case-insensitive fallback
        try:
            for f in search_dir.iterdir():
                if f.name.lower() == p.name.lower():
                    return True
        except Exception:
            pass
        return False

    # ------------------------------------------------------------------ #
    # Colour helpers
    # ------------------------------------------------------------------ #

    @staticmethod
    def _pick_fill(match_type: str, similarity: float, alt: bool):
        from openpyxl.styles import PatternFill

        if match_type == "exact":
            return PatternFill("solid", fgColor=_EXACT_FILL)
        elif match_type == "exact_subsequence":
            return PatternFill("solid", fgColor=_EXACT_SUB_FILL)
        elif match_type == "approximate":
            if similarity >= 0.95:
                return PatternFill("solid", fgColor=_APPROX_HIGH_FILL)
            else:
                return PatternFill("solid", fgColor=_APPROX_MED_FILL)
        else:
            return PatternFill("solid", fgColor=_UNMATCHED_FILL)
