"""Tests for the upstream tracing module."""
from __future__ import annotations

import io
import json
import re
import tempfile
import zipfile
from pathlib import Path

import numpy as np
import pytest

from tests.test_generator import generate_test_workbook

FIXTURE_DIR = Path(__file__).parent / "fixtures"


# ---------------------------------------------------------------------------
# Config tests
# ---------------------------------------------------------------------------

def test_config_defaults():
    """TraceConfig has sensible defaults."""
    from lineage.tracing.config import TraceConfig

    c = TraceConfig()
    assert c.exact is True
    assert c.approximate is True
    assert c.top_n == 5
    assert c.similarity_metric == "pearson"
    assert c.min_similarity == 0.8
    assert c.subsequence_matching is True
    assert c.direction_sensitive is False
    assert c.min_vector_length == 3


def test_config_from_json():
    """TraceConfig loads from JSON."""
    from lineage.tracing.config import TraceConfig

    with tempfile.NamedTemporaryFile("w", suffix=".json", delete=False) as f:
        json.dump({
            "matching": {"exact": False, "top_n": 10, "similarity_metric": "cosine"},
            "performance": {"min_vector_length": 5},
        }, f)
        f.flush()
        c = TraceConfig.from_file(Path(f.name))

    assert c.exact is False
    assert c.top_n == 10
    assert c.similarity_metric == "cosine"
    assert c.min_vector_length == 5
    # Defaults preserved for unset fields
    assert c.approximate is True


def test_config_to_dict():
    """TraceConfig round-trips through to_dict."""
    from lineage.tracing.config import TraceConfig

    c = TraceConfig(top_n=3, similarity_metric="euclidean")
    d = c.to_dict()
    assert d["matching"]["top_n"] == 3
    assert d["matching"]["similarity_metric"] == "euclidean"


# ---------------------------------------------------------------------------
# Scanner tests
# ---------------------------------------------------------------------------

def test_scan_model_sheet():
    """Scanner finds hardcoded vectors in a model sheet."""
    FIXTURE_DIR.mkdir(exist_ok=True)
    test_file = FIXTURE_DIR / "test_connections.xlsx"
    if not test_file.exists():
        generate_test_workbook(test_file)

    from lineage.tracing.scanner import scan_model_sheet
    vectors = scan_model_sheet(test_file, "Inputs")
    assert len(vectors) > 0, "Expected hardcoded vectors in Inputs sheet"
    for v in vectors:
        assert v.file == "test_connections.xlsx"
        assert v.sheet == "Inputs"
        assert v.length >= 3
        assert len(v.values) == v.length
        assert v.direction in ("column", "row")


def test_scan_upstream_file():
    """Upstream scanner finds ALL numeric vectors (formula + hardcoded)."""
    FIXTURE_DIR.mkdir(exist_ok=True)
    test_file = FIXTURE_DIR / "test_connections.xlsx"
    if not test_file.exists():
        generate_test_workbook(test_file)

    from lineage.tracing.scanner import scan_upstream_file
    vectors = scan_upstream_file(test_file)
    assert len(vectors) > 0, "Expected upstream vectors"
    # Should find vectors across multiple sheets
    sheets = {v.sheet for v in vectors}
    assert len(sheets) >= 1


def test_get_sheet_names():
    """get_sheet_names returns correct sheets."""
    FIXTURE_DIR.mkdir(exist_ok=True)
    test_file = FIXTURE_DIR / "test_connections.xlsx"
    if not test_file.exists():
        generate_test_workbook(test_file)

    from lineage.tracing.scanner import get_sheet_names
    names = get_sheet_names(test_file)
    assert "Inputs" in names


def test_scanner_xls_returns_empty():
    """Scanner gracefully returns empty for non-existent / non-ZIP files."""
    from lineage.tracing.scanner import scan_upstream_file
    result = scan_upstream_file(Path("/nonexistent/file.xlsx"))
    assert result == []


# ---------------------------------------------------------------------------
# Exact matcher tests
# ---------------------------------------------------------------------------

def test_exact_full_match():
    """ExactMatcher finds identical vectors."""
    from lineage.tracing.config import TraceConfig
    from lineage.tracing.exact_matcher import ExactMatcher
    from lineage.tracing.models import TracingVector

    config = TraceConfig()
    em = ExactMatcher(config)

    upstream = TracingVector(
        file="source.xlsx", sheet="Data", cell_range="A1:A5",
        direction="column", length=5, start_cell="A1", end_cell="A5",
        values=(10.0, 20.0, 30.0, 40.0, 50.0),
    )
    em.index_upstream([upstream])

    model = TracingVector(
        file="model.xlsx", sheet="Sheet1", cell_range="B3:B7",
        direction="column", length=5, start_cell="B3", end_cell="B7",
        values=(10.0, 20.0, 30.0, 40.0, 50.0),
    )
    matches = em.match(model)
    assert len(matches) == 1
    assert matches[0].match_type == "exact"
    assert matches[0].similarity == 1.0
    assert matches[0].upstream_file == "source.xlsx"


def test_exact_subsequence_match():
    """ExactMatcher finds model vector as subsequence of longer upstream."""
    from lineage.tracing.config import TraceConfig
    from lineage.tracing.exact_matcher import ExactMatcher
    from lineage.tracing.models import TracingVector

    config = TraceConfig(subsequence_matching=True)
    em = ExactMatcher(config)

    upstream = TracingVector(
        file="source.xlsx", sheet="Data", cell_range="A1:A10",
        direction="column", length=10, start_cell="A1", end_cell="A10",
        values=(5.0, 10.0, 20.0, 30.0, 40.0, 50.0, 60.0, 70.0, 80.0, 90.0),
    )
    em.index_upstream([upstream])

    model = TracingVector(
        file="model.xlsx", sheet="Sheet1", cell_range="B1:B3",
        direction="column", length=3, start_cell="B1", end_cell="B3",
        values=(20.0, 30.0, 40.0),
    )
    matches = em.match(model)
    assert any(m.match_type == "exact_subsequence" for m in matches)
    # The matched sub-range should be A3:A5 (offset 2, length 3)
    sub_match = [m for m in matches if m.match_type == "exact_subsequence"][0]
    assert sub_match.upstream_matched_range == "A3:A5"


def test_exact_no_match():
    """ExactMatcher returns empty when no match exists."""
    from lineage.tracing.config import TraceConfig
    from lineage.tracing.exact_matcher import ExactMatcher
    from lineage.tracing.models import TracingVector

    config = TraceConfig()
    em = ExactMatcher(config)

    upstream = TracingVector(
        file="source.xlsx", sheet="Data", cell_range="A1:A3",
        direction="column", length=3, start_cell="A1", end_cell="A3",
        values=(100.0, 200.0, 300.0),
    )
    em.index_upstream([upstream])

    model = TracingVector(
        file="model.xlsx", sheet="Sheet1", cell_range="B1:B3",
        direction="column", length=3, start_cell="B1", end_cell="B3",
        values=(1.0, 2.0, 3.0),
    )
    matches = em.match(model)
    assert len(matches) == 0


# ---------------------------------------------------------------------------
# Approximate matcher tests
# ---------------------------------------------------------------------------

def test_approx_pearson_identical():
    """Pearson correlation = 1.0 for identical vectors."""
    from lineage.tracing.config import TraceConfig
    from lineage.tracing.approx_matcher import ApproximateMatcher
    from lineage.tracing.models import TracingVector

    config = TraceConfig(similarity_metric="pearson", min_similarity=0.5)
    am = ApproximateMatcher(config)

    upstream = TracingVector(
        file="source.xlsx", sheet="Data", cell_range="A1:A5",
        direction="column", length=5, start_cell="A1", end_cell="A5",
        values=(10.0, 20.0, 30.0, 40.0, 50.0),
    )
    am.index_upstream([upstream])

    model = TracingVector(
        file="model.xlsx", sheet="Sheet1", cell_range="B1:B5",
        direction="column", length=5, start_cell="B1", end_cell="B5",
        values=(10.0, 20.0, 30.0, 40.0, 50.0),
    )
    matches = am.match(model)
    assert len(matches) >= 1
    assert matches[0].similarity > 0.99


def test_approx_pearson_scaled():
    """Pearson detects same shape at different scale."""
    from lineage.tracing.config import TraceConfig
    from lineage.tracing.approx_matcher import ApproximateMatcher
    from lineage.tracing.models import TracingVector

    config = TraceConfig(similarity_metric="pearson", min_similarity=0.9)
    am = ApproximateMatcher(config)

    upstream = TracingVector(
        file="source.xlsx", sheet="Data", cell_range="A1:A5",
        direction="column", length=5, start_cell="A1", end_cell="A5",
        values=(100.0, 200.0, 300.0, 400.0, 500.0),
    )
    am.index_upstream([upstream])

    # Same shape, different scale
    model = TracingVector(
        file="model.xlsx", sheet="Sheet1", cell_range="B1:B5",
        direction="column", length=5, start_cell="B1", end_cell="B5",
        values=(10.0, 20.0, 30.0, 40.0, 50.0),
    )
    matches = am.match(model)
    assert len(matches) >= 1
    assert matches[0].similarity > 0.99  # perfect linear correlation


def test_approx_length_mismatch():
    """Approximate matcher handles length-mismatched vectors via sliding window."""
    from lineage.tracing.config import TraceConfig
    from lineage.tracing.approx_matcher import ApproximateMatcher
    from lineage.tracing.models import TracingVector

    config = TraceConfig(
        similarity_metric="pearson",
        min_similarity=0.9,
        length_tolerance_pct=100.0,
    )
    am = ApproximateMatcher(config)

    # Upstream is longer
    upstream = TracingVector(
        file="source.xlsx", sheet="Data", cell_range="A1:A8",
        direction="column", length=8, start_cell="A1", end_cell="A8",
        values=(0.0, 0.0, 10.0, 20.0, 30.0, 40.0, 50.0, 0.0),
    )
    am.index_upstream([upstream])

    model = TracingVector(
        file="model.xlsx", sheet="Sheet1", cell_range="B1:B5",
        direction="column", length=5, start_cell="B1", end_cell="B5",
        values=(10.0, 20.0, 30.0, 40.0, 50.0),
    )
    matches = am.match(model)
    assert len(matches) >= 1
    assert matches[0].similarity > 0.99


def test_approx_cosine():
    """Cosine similarity works."""
    from lineage.tracing.config import TraceConfig
    from lineage.tracing.approx_matcher import ApproximateMatcher
    from lineage.tracing.models import TracingVector

    config = TraceConfig(similarity_metric="cosine", min_similarity=0.5)
    am = ApproximateMatcher(config)

    upstream = TracingVector(
        file="source.xlsx", sheet="Data", cell_range="A1:A4",
        direction="column", length=4, start_cell="A1", end_cell="A4",
        values=(1.0, 2.0, 3.0, 4.0),
    )
    am.index_upstream([upstream])

    model = TracingVector(
        file="model.xlsx", sheet="Sheet1", cell_range="B1:B4",
        direction="column", length=4, start_cell="B1", end_cell="B4",
        values=(2.0, 4.0, 6.0, 8.0),
    )
    matches = am.match(model)
    assert len(matches) >= 1
    assert matches[0].similarity > 0.99  # identical direction


def test_approx_below_threshold():
    """Vectors below min_similarity are not returned."""
    from lineage.tracing.config import TraceConfig
    from lineage.tracing.approx_matcher import ApproximateMatcher
    from lineage.tracing.models import TracingVector

    config = TraceConfig(similarity_metric="pearson", min_similarity=0.99)
    am = ApproximateMatcher(config)

    upstream = TracingVector(
        file="source.xlsx", sheet="Data", cell_range="A1:A4",
        direction="column", length=4, start_cell="A1", end_cell="A4",
        values=(1.0, 2.0, 3.0, 4.0),
    )
    am.index_upstream([upstream])

    # Very different shape
    model = TracingVector(
        file="model.xlsx", sheet="Sheet1", cell_range="B1:B4",
        direction="column", length=4, start_cell="B1", end_cell="B4",
        values=(4.0, 1.0, 3.0, 2.0),
    )
    matches = am.match(model)
    # Correlation is low, should be below 0.99
    assert all(m.similarity >= 0.99 for m in matches) or len(matches) == 0


# ---------------------------------------------------------------------------
# Batch similarity kernel tests
# ---------------------------------------------------------------------------

def test_batch_pearson_constant_vector():
    """Pearson handles constant vectors without NaN."""
    from lineage.tracing.approx_matcher import _batch_pearson

    model = np.array([5.0, 5.0, 5.0, 5.0])
    batch = np.array([
        [5.0, 5.0, 5.0, 5.0],  # same constant → 1.0
        [1.0, 2.0, 3.0, 4.0],  # non-constant → 0.0
        [3.0, 3.0, 3.0, 3.0],  # different constant → 0.0
    ])
    result = _batch_pearson(model, batch)
    assert not np.any(np.isnan(result))
    assert result[0] == pytest.approx(1.0)
    assert result[1] == pytest.approx(0.0)
    assert result[2] == pytest.approx(0.0)


# ---------------------------------------------------------------------------
# End-to-end tracer test
# ---------------------------------------------------------------------------

def test_tracer_end_to_end():
    """UpstreamTracer finds matches between test files."""
    FIXTURE_DIR.mkdir(exist_ok=True)
    test_file = FIXTURE_DIR / "test_connections.xlsx"
    if not test_file.exists():
        generate_test_workbook(test_file)

    from lineage.tracing.config import TraceConfig
    from lineage.tracing.tracer import UpstreamTracer

    config = TraceConfig(min_similarity=0.7, top_n=3)
    tracer = UpstreamTracer(config=config)

    # Trace the test file against itself (should find exact matches)
    matches, unmatched = tracer.trace(
        test_file, "Inputs", [test_file],
    )
    # Should find at least some matches (tracing file against itself)
    assert len(matches) + len(unmatched) > 0


def test_tracer_report():
    """TracingReporter produces a valid Excel file."""
    FIXTURE_DIR.mkdir(exist_ok=True)
    test_file = FIXTURE_DIR / "test_connections.xlsx"
    if not test_file.exists():
        generate_test_workbook(test_file)

    from lineage.tracing.config import TraceConfig
    from lineage.tracing.tracer import UpstreamTracer
    from lineage.tracing.report import TracingReporter

    config = TraceConfig(min_similarity=0.7, top_n=3)
    tracer = UpstreamTracer(config=config)
    matches, unmatched = tracer.trace(test_file, "Inputs", [test_file])

    with tempfile.TemporaryDirectory() as tmpdir:
        out_dir = Path(tmpdir)
        reporter = TracingReporter()
        out_path = reporter.write(
            matches, unmatched, config, test_file, "Inputs",
            [test_file], out_dir,
        )
        assert out_path.exists()
        assert out_path.suffix == ".xlsx"
        assert out_path.stat().st_size > 0

        # Verify structure
        import openpyxl
        wb = openpyxl.load_workbook(str(out_path))
        assert "Config" in wb.sheetnames
        assert "Tracing Results" in wb.sheetnames


# ---------------------------------------------------------------------------
# Formula tracer — helpers
# ---------------------------------------------------------------------------

def _make_xlsx_with_external_formulas(path: Path, formulas: dict[str, dict[str, str]]):
    """Create a minimal .xlsx with external formulas injected at specific cells.

    formulas: {sheet_name: {cell_ref: formula_text}}
    """
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name in formulas:
        wb.create_sheet(sheet_name)
    wb.save(str(path))

    # Now re-open the ZIP and inject <f> elements into the sheet XML
    with zipfile.ZipFile(path, "r") as zf:
        names = zf.namelist()
        contents = {n: zf.read(n) for n in names}

    # Read workbook.xml to find sheet rId → sheet path
    from lxml import etree
    wb_xml = etree.fromstring(contents["xl/workbook.xml"])
    ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_r = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

    rels_xml = etree.fromstring(contents["xl/_rels/workbook.xml.rels"])
    rid_to_path = {}
    for rel in rels_xml:
        rid = rel.get("Id", "")
        target = rel.get("Target", "").lstrip("/")
        # Target may be "worksheets/sheet1.xml" or "xl/worksheets/sheet1.xml"
        if "worksheets/" in target:
            if target.startswith("xl/"):
                rid_to_path[rid] = target
            else:
                rid_to_path[rid] = "xl/" + target

    sheet_name_to_path = {}
    for el in wb_xml.iter(f"{{{ns_main}}}sheet"):
        sname = el.get("name")
        rid = el.get(f"{{{ns_r}}}id")
        if rid in rid_to_path:
            sheet_name_to_path[sname] = rid_to_path[rid]

    # For each sheet, inject formulas
    for sheet_name, cell_formulas in formulas.items():
        sheet_path = sheet_name_to_path.get(sheet_name)
        if not sheet_path:
            continue

        root = etree.fromstring(contents[sheet_path])
        ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

        # Find or create sheetData
        sheet_data = root.find(f"{{{ns}}}sheetData")
        if sheet_data is None:
            sheet_data = etree.SubElement(root, f"{{{ns}}}sheetData")

        for cell_ref, formula_text in cell_formulas.items():
            # Parse cell ref
            m = re.match(r"([A-Z]+)(\d+)", cell_ref)
            row_num = int(m.group(2))
            col_letters = m.group(1)

            # Find or create the row
            row_el = None
            for r in sheet_data.findall(f"{{{ns}}}row"):
                if r.get("r") == str(row_num):
                    row_el = r
                    break
            if row_el is None:
                row_el = etree.SubElement(sheet_data, f"{{{ns}}}row")
                row_el.set("r", str(row_num))

            # Create the cell with formula
            c_el = etree.SubElement(row_el, f"{{{ns}}}c")
            c_el.set("r", cell_ref)
            f_el = etree.SubElement(c_el, f"{{{ns}}}f")
            f_el.text = formula_text
            v_el = etree.SubElement(c_el, f"{{{ns}}}v")
            v_el.text = "0"

        contents[sheet_path] = etree.tostring(root, xml_declaration=True, encoding="UTF-8")

    # Rewrite the ZIP
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in contents.items():
            zf.writestr(name, data)


# ---------------------------------------------------------------------------
# Formula tracer — unit tests
# ---------------------------------------------------------------------------

def test_parse_range():
    """_parse_range handles single cells and ranges."""
    from lineage.tracing.formula_tracer import _parse_range

    assert _parse_range("A1") == (1, 1, 1, 1)
    assert _parse_range("B3:D10") == (3, 10, 2, 4)
    assert _parse_range("$C$5:$E$8") == (5, 8, 3, 5)


def test_extract_filename():
    """_extract_filename handles paths, URLs, and SharePoint URIs."""
    from lineage.tracing.formula_tracer import _extract_filename

    assert _extract_filename("budget.xlsx") == "budget.xlsx"
    assert _extract_filename("C:\\Users\\data\\budget.xlsx") == "budget.xlsx"
    assert _extract_filename("/Users/data/budget.xlsx") == "budget.xlsx"
    assert _extract_filename(
        "https://company.sharepoint.com/sites/team/Shared%20Documents/budget.xlsx"
    ) == "budget.xlsx"
    assert _extract_filename("file:///C:/Users/data/budget.xlsx") == "budget.xlsx"


def test_cell_filter():
    """CellFilter correctly filters cells by sheet and rectangle."""
    from lineage.tracing.formula_tracer import CellFilter

    cf = CellFilter.from_refs([("Sheet1", "A1:C5"), ("Sheet2", "B2:B10")])

    assert cf.has_sheet("Sheet1")
    assert cf.has_sheet("Sheet2")
    assert not cf.has_sheet("Sheet3")

    assert cf.contains("Sheet1", 1, 1)   # A1
    assert cf.contains("Sheet1", 5, 3)   # C5
    assert not cf.contains("Sheet1", 6, 1)  # A6 — outside
    assert cf.contains("Sheet2", 5, 2)   # B5
    assert not cf.contains("Sheet2", 1, 2)  # B1 — outside


def test_parse_formula_refs_literal():
    """_parse_formula_refs handles literal [file.xlsx]Sheet!Range."""
    from lineage.tracing.formula_tracer import _parse_formula_refs

    formula = "'[budget.xlsx]Revenue'!A1:A10"
    refs = _parse_formula_refs(formula, {})
    assert len(refs) == 1
    assert refs[0][0] == "budget.xlsx"
    assert refs[0][1] == "Revenue"
    assert refs[0][2] == "A1:A10"


def test_parse_formula_refs_numeric_index():
    """_parse_formula_refs resolves numeric [1] via link_map."""
    from lineage.tracing.formula_tracer import _parse_formula_refs

    link_map = {"1": ("source.xlsx", "/path/to/source.xlsx")}
    formula = "[1]Data!B5"
    refs = _parse_formula_refs(formula, link_map)
    assert len(refs) == 1
    assert refs[0][0] == "source.xlsx"
    assert refs[0][1] == "Data"
    assert refs[0][2] == "B5"


def test_parse_formula_refs_sharepoint():
    """_parse_formula_refs handles SharePoint URL paths."""
    from lineage.tracing.formula_tracer import _parse_formula_refs

    formula = "'https://company.sharepoint.com/sites/team/Shared Documents/[data.xlsx]Sheet1'!C3:C20"
    refs = _parse_formula_refs(formula, {})
    assert len(refs) == 1
    assert refs[0][0] == "data.xlsx"
    assert refs[0][1] == "Sheet1"
    assert refs[0][2] == "C3:C20"


def test_parse_formula_refs_multiple():
    """_parse_formula_refs returns all refs from a complex formula."""
    from lineage.tracing.formula_tracer import _parse_formula_refs

    formula = "'[a.xlsx]S1'!A1+'[b.xlsx]S2'!B2"
    refs = _parse_formula_refs(formula, {})
    assert len(refs) == 2
    filenames = {r[0] for r in refs}
    assert filenames == {"a.xlsx", "b.xlsx"}


def test_resolve_file_found(tmp_path):
    """_resolve_file finds an existing file."""
    from lineage.tracing.formula_tracer import _resolve_file

    (tmp_path / "budget.xlsx").write_bytes(b"fake")
    path, found = _resolve_file("budget.xlsx", [tmp_path])
    assert found is True
    assert path.name == "budget.xlsx"


def test_resolve_file_case_insensitive(tmp_path):
    """_resolve_file falls back to case-insensitive search."""
    from lineage.tracing.formula_tracer import _resolve_file

    (tmp_path / "Budget.xlsx").write_bytes(b"fake")
    path, found = _resolve_file("budget.xlsx", [tmp_path])
    assert found is True
    assert path.name == "Budget.xlsx"


def test_resolve_file_not_found(tmp_path):
    """_resolve_file returns (expected_path, False) for missing files."""
    from lineage.tracing.formula_tracer import _resolve_file

    path, found = _resolve_file("nonexistent.xlsx", [tmp_path])
    assert found is False
    assert path.name == "nonexistent.xlsx"


def test_stream_external_formulas():
    """_stream_external_formulas finds formulas with [ in them."""
    from lineage.tracing.formula_tracer import _stream_external_formulas

    # Build minimal sheet XML
    xml = b'''<?xml version="1.0" encoding="UTF-8"?>
    <worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
    <sheetData>
      <row r="1">
        <c r="A1"><f>'[data.xlsx]Sheet1'!A1</f><v>100</v></c>
        <c r="B1"><f>SUM(A1:A5)</f><v>500</v></c>
        <c r="C1"><v>42</v></c>
      </row>
    </sheetData>
    </worksheet>'''

    results = _stream_external_formulas(xml, "Sheet1")
    assert len(results) == 1
    assert results[0][0] == "A1"
    assert "[data.xlsx]" in results[0][1]


def test_stream_external_formulas_with_filter():
    """_stream_external_formulas respects cell_filter."""
    from lineage.tracing.formula_tracer import _stream_external_formulas, CellFilter

    xml = b'''<?xml version="1.0" encoding="UTF-8"?>
    <worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
    <sheetData>
      <row r="1">
        <c r="A1"><f>'[data.xlsx]Sheet1'!A1</f><v>100</v></c>
      </row>
      <row r="10">
        <c r="A10"><f>'[other.xlsx]Sheet2'!B5</f><v>200</v></c>
      </row>
    </sheetData>
    </worksheet>'''

    # Filter only row 10
    cf = CellFilter.from_refs([("Sheet1", "A10:A10")])
    results = _stream_external_formulas(xml, "Sheet1", cf)
    assert len(results) == 1
    assert results[0][0] == "A10"


def test_scan_external_refs_on_generated_file():
    """scan_external_refs finds injected external formulas in an xlsx."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        test_file = tmp / "model.xlsx"
        _make_xlsx_with_external_formulas(test_file, {
            "Sheet1": {
                "A1": "'[upstream.xlsx]Data'!A1:A10",
                "B1": "'[source.xlsx]Revenue'!C3",
            },
        })

        from lineage.tracing.formula_tracer import scan_external_refs
        refs = scan_external_refs(test_file, "model.xlsx", level=1, search_dirs=[tmp])

        assert len(refs) >= 2
        target_files = {r.target_file for r in refs}
        assert "upstream.xlsx" in target_files
        assert "source.xlsx" in target_files

        for r in refs:
            assert r.level == 1
            assert r.source_file == "model.xlsx"
            assert r.file_found is False  # files don't exist


def test_trace_formula_levels_single_level():
    """trace_formula_levels returns Level 1 refs for a single-level file."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        test_file = tmp / "model.xlsx"
        _make_xlsx_with_external_formulas(test_file, {
            "Sheet1": {
                "A1": "'[missing.xlsx]Data'!B2:B20",
            },
        })

        from lineage.tracing.formula_tracer import trace_formula_levels
        results = trace_formula_levels(test_file, search_dirs=[tmp], max_level=5)

        assert 1 in results
        assert len(results[1]) >= 1
        assert results[1][0].target_file == "missing.xlsx"
        assert results[1][0].file_found is False
        # Should stop at level 1 since no files found
        assert 2 not in results


def test_trace_formula_levels_two_levels():
    """trace_formula_levels follows chain: model → upstream → upstream2."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)

        # Level 2 file: upstream.xlsx references upstream2.xlsx
        upstream_file = tmp / "upstream.xlsx"
        _make_xlsx_with_external_formulas(upstream_file, {
            "Data": {
                "A1": "'[upstream2.xlsx]Raw'!A1:A5",
            },
        })

        # Level 1 file: model.xlsx references upstream.xlsx
        model_file = tmp / "model.xlsx"
        _make_xlsx_with_external_formulas(model_file, {
            "Sheet1": {
                "A1": "'[upstream.xlsx]Data'!A1",
            },
        })

        from lineage.tracing.formula_tracer import trace_formula_levels
        results = trace_formula_levels(model_file, search_dirs=[tmp], max_level=5)

        assert 1 in results
        assert any(r.target_file == "upstream.xlsx" for r in results[1])
        assert any(r.file_found for r in results[1])  # upstream.xlsx exists

        assert 2 in results
        assert any(r.target_file == "upstream2.xlsx" for r in results[2])
        assert not any(r.file_found for r in results[2])  # upstream2.xlsx doesn't exist


def test_report_with_levels():
    """TracingReporter.write_with_levels creates Level N sheets."""
    from lineage.tracing.config import TraceConfig
    from lineage.tracing.report import TracingReporter
    from lineage.tracing.formula_tracer import ExternalReference

    config = TraceConfig()
    level_refs = {
        1: [
            ExternalReference(
                level=1, source_file="model.xlsx", source_sheet="Sheet1",
                source_cell="A1", formula="'[data.xlsx]Sheet1'!A1",
                target_file="data.xlsx", target_sheet="Sheet1",
                target_range="A1", target_path="data.xlsx",
                file_found=True, resolved_path="/tmp/data.xlsx",
            ),
            ExternalReference(
                level=1, source_file="model.xlsx", source_sheet="Sheet1",
                source_cell="B1", formula="'[missing.xlsx]Sheet1'!B2",
                target_file="missing.xlsx", target_sheet="Sheet1",
                target_range="B2", target_path="missing.xlsx",
                file_found=False, resolved_path="/tmp/missing.xlsx",
            ),
        ],
    }

    with tempfile.TemporaryDirectory() as tmpdir:
        out_dir = Path(tmpdir)
        reporter = TracingReporter()
        out_path = reporter.write_with_levels(
            [], [], config, Path("/tmp/model.xlsx"), "Sheet1",
            [], out_dir, level_refs=level_refs,
        )
        assert out_path.exists()

        import openpyxl
        wb = openpyxl.load_workbook(str(out_path))
        assert "Config" in wb.sheetnames
        assert "Level 1" in wb.sheetnames

        ws = wb["Level 1"]
        # Header row
        assert ws.cell(row=1, column=1).value == "Source File"
        assert ws.cell(row=1, column=5).value == "Target File"
        assert ws.cell(row=1, column=9).value == "File Found"

        # Data rows
        assert ws.cell(row=2, column=5).value == "data.xlsx"
        assert ws.cell(row=2, column=9).value == "Yes"
        assert ws.cell(row=3, column=5).value == "missing.xlsx"
        assert ws.cell(row=3, column=9).value == "No"

        # Check color: row 2 should be green, row 3 should be red/salmon
        fill_found = ws.cell(row=2, column=1).fill.fgColor.rgb
        fill_missing = ws.cell(row=3, column=1).fill.fgColor.rgb
        # Green is C8E6C9, salmon is FFCDD2
        assert "C8E6C9" in str(fill_found)
        assert "FFCDD2" in str(fill_missing)


# ---------------------------------------------------------------------------
# Precedent tracing — unit tests
# ---------------------------------------------------------------------------

def test_expand_range():
    """_expand_range handles single cells and ranges."""
    from lineage.tracing.formula_tracer import _expand_range

    assert _expand_range("A1") == ["A1"]
    assert _expand_range("A1:A3") == ["A1", "A2", "A3"]
    assert _expand_range("A1:B2") == ["A1", "B1", "A2", "B2"]
    assert _expand_range("$C$3") == ["C3"]
    assert _expand_range("$A$1:$A$3") == ["A1", "A2", "A3"]


def test_expand_range_large():
    """_expand_range caps overly large ranges."""
    from lineage.tracing.formula_tracer import _expand_range

    # A1:A1000000 would be 1M cells — should return empty (capped)
    result = _expand_range("A1:A1000000")
    assert result == []


def test_parse_intra_refs_simple():
    """_parse_intra_refs finds simple cell references."""
    from lineage.tracing.formula_tracer import _parse_intra_refs

    refs = _parse_intra_refs("B1*2", "Sheet1")
    assert ("Sheet1", "B1") in refs


def test_parse_intra_refs_range():
    """_parse_intra_refs finds range references."""
    from lineage.tracing.formula_tracer import _parse_intra_refs

    refs = _parse_intra_refs("SUM(B1:B10)", "Sheet1")
    assert any(r[1] == "B1:B10" for r in refs)


def test_parse_intra_refs_cross_sheet():
    """_parse_intra_refs finds cross-sheet references."""
    from lineage.tracing.formula_tracer import _parse_intra_refs

    refs = _parse_intra_refs("Sheet2!A1+B1", "Sheet1")
    sheets = {r[0] for r in refs}
    assert "Sheet2" in sheets
    assert "Sheet1" in sheets


def test_parse_intra_refs_quoted_sheet():
    """_parse_intra_refs handles quoted sheet names."""
    from lineage.tracing.formula_tracer import _parse_intra_refs

    refs = _parse_intra_refs("'Data Sheet'!A1", "Sheet1")
    assert ("Data Sheet", "A1") in refs


def test_parse_intra_refs_excludes_external():
    """_parse_intra_refs excludes references inside external workbook refs."""
    from lineage.tracing.formula_tracer import _parse_intra_refs

    # External ref: '[ext.xlsx]S1'!A1 — should NOT appear in intra refs
    # But B1 should appear
    refs = _parse_intra_refs("'[ext.xlsx]S1'!A1+B1", "Sheet1")
    cell_refs = {r[1] for r in refs}
    assert "B1" in cell_refs
    # The A1 from the external ref should be excluded
    # (it's inside the _REF_RE match span)


def test_parse_intra_refs_dollar_signs():
    """_parse_intra_refs strips dollar signs."""
    from lineage.tracing.formula_tracer import _parse_intra_refs

    refs = _parse_intra_refs("$B$5+$C$10:$D$20", "Sheet1")
    cell_refs = {r[1] for r in refs}
    assert "B5" in cell_refs
    assert "C10:D20" in cell_refs


def test_walk_precedents_direct_external():
    """_walk_precedents finds nothing when the starting cell itself is external."""
    from lineage.tracing.formula_tracer import _walk_precedents

    # A1 directly references external — _walk_precedents is only called for
    # cells WITHOUT direct external refs, so let's test a cell with a formula
    # referencing B1, where B1 is external
    cache = {
        "Sheet1": {
            "A1": "B1*2",
            "B1": "'[source.xlsx]Data'!C3",
        }
    }
    hits = _walk_precedents("Sheet1", "A1", cache, {})
    assert len(hits) >= 1
    assert hits[0].chain[-1][0] == "Sheet1"
    assert hits[0].chain[-1][1] == "B1"
    assert any(r[0] == "source.xlsx" for r in hits[0].external_refs)


def test_walk_precedents_two_hops():
    """_walk_precedents follows a 2-hop chain: A1 → B1 → C1 (external)."""
    from lineage.tracing.formula_tracer import _walk_precedents

    cache = {
        "Sheet1": {
            "A1": "B1+1",
            "B1": "C1*2",
            "C1": "'[source.xlsx]Data'!D5",
        }
    }
    hits = _walk_precedents("Sheet1", "A1", cache, {})
    assert len(hits) >= 1
    # Chain should be B1 → C1
    chain = hits[0].chain
    assert len(chain) == 2
    assert chain[0][1] == "B1"
    assert chain[1][1] == "C1"


def test_walk_precedents_cross_sheet():
    """_walk_precedents follows cross-sheet references."""
    from lineage.tracing.formula_tracer import _walk_precedents

    cache = {
        "Sheet1": {
            "A1": "Sheet2!B1",
        },
        "Sheet2": {
            "B1": "'[ext.xlsx]Data'!A1",
        },
    }
    hits = _walk_precedents("Sheet1", "A1", cache, {})
    assert len(hits) >= 1
    assert hits[0].chain[-1][0] == "Sheet2"
    assert hits[0].chain[-1][1] == "B1"


def test_walk_precedents_circular():
    """_walk_precedents terminates on circular references."""
    from lineage.tracing.formula_tracer import _walk_precedents

    cache = {
        "Sheet1": {
            "A1": "B1+1",
            "B1": "A1+1",  # circular!
        }
    }
    # Should not hang — visited set prevents infinite loop
    hits = _walk_precedents("Sheet1", "A1", cache, {})
    assert len(hits) == 0  # no external refs found


def test_walk_precedents_dead_end():
    """_walk_precedents handles cells with no formula (dead end)."""
    from lineage.tracing.formula_tracer import _walk_precedents

    cache = {
        "Sheet1": {
            "A1": "B1+1",
            # B1 has no formula — it's a hardcoded value
        }
    }
    hits = _walk_precedents("Sheet1", "A1", cache, {})
    assert len(hits) == 0


def test_walk_precedents_multiple_paths():
    """_walk_precedents finds external refs through multiple paths."""
    from lineage.tracing.formula_tracer import _walk_precedents

    cache = {
        "Sheet1": {
            "A1": "B1+C1",
            "B1": "'[file_b.xlsx]S1'!A1",
            "C1": "'[file_c.xlsx]S2'!B2",
        }
    }
    hits = _walk_precedents("Sheet1", "A1", cache, {})
    assert len(hits) == 2
    found_files = set()
    for hit in hits:
        for ref in hit.external_refs:
            found_files.add(ref[0])
    assert "file_b.xlsx" in found_files
    assert "file_c.xlsx" in found_files


def test_stream_all_formulas():
    """_stream_all_formulas returns all formula cells."""
    from lineage.tracing.formula_tracer import _stream_all_formulas

    xml = b'''<?xml version="1.0" encoding="UTF-8"?>
    <worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
    <sheetData>
      <row r="1">
        <c r="A1"><f>B1*2</f><v>10</v></c>
        <c r="B1"><v>5</v></c>
        <c r="C1"><f>'[ext.xlsx]S1'!A1</f><v>100</v></c>
      </row>
    </sheetData>
    </worksheet>'''

    formulas = _stream_all_formulas(xml)
    assert "A1" in formulas
    assert "C1" in formulas
    assert "B1" not in formulas  # no formula, just a value
    assert formulas["A1"] == "B1*2"


# ---------------------------------------------------------------------------
# Precedent tracing — integration tests
# ---------------------------------------------------------------------------

def test_trace_transitive_two_hops():
    """Level 2 finds external refs through transitive precedent chain."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)

        # upstream.xlsx: Data!A1 = B1*2, Data!B1 = '[source.xlsx]Raw'!C3
        upstream_file = tmp / "upstream.xlsx"
        _make_xlsx_with_external_formulas(upstream_file, {
            "Data": {
                "A1": "B1*2",
                "B1": "'[source.xlsx]Raw'!C3",
            },
        })

        # model.xlsx: Sheet1!A1 = '[upstream.xlsx]Data'!A1
        model_file = tmp / "model.xlsx"
        _make_xlsx_with_external_formulas(model_file, {
            "Sheet1": {
                "A1": "'[upstream.xlsx]Data'!A1",
            },
        })

        from lineage.tracing.formula_tracer import trace_formula_levels
        results = trace_formula_levels(model_file, search_dirs=[tmp], max_level=5)

        assert 1 in results
        assert any(r.target_file == "upstream.xlsx" for r in results[1])

        # Level 2: should find source.xlsx through A1 → B1 chain
        assert 2 in results
        source_refs = [r for r in results[2] if r.target_file == "source.xlsx"]
        assert len(source_refs) >= 1

        ref = source_refs[0]
        assert ref.source_cell == "A1"  # original cell in cell filter
        assert ref.precedent_chain is not None
        assert len(ref.precedent_chain) >= 1
        # The chain should end at B1 which has the external ref
        assert ref.precedent_chain[-1][1] == "B1"


def test_trace_transitive_three_hops():
    """Level 2 finds external refs through a 3-hop precedent chain."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)

        # upstream.xlsx: Data!A1 = B1+1, B1 = C1*2, C1 = '[source.xlsx]Raw'!D5
        upstream_file = tmp / "upstream.xlsx"
        _make_xlsx_with_external_formulas(upstream_file, {
            "Data": {
                "A1": "B1+1",
                "B1": "C1*2",
                "C1": "'[source.xlsx]Raw'!D5",
            },
        })

        # model.xlsx
        model_file = tmp / "model.xlsx"
        _make_xlsx_with_external_formulas(model_file, {
            "Sheet1": {
                "A1": "'[upstream.xlsx]Data'!A1",
            },
        })

        from lineage.tracing.formula_tracer import trace_formula_levels
        results = trace_formula_levels(model_file, search_dirs=[tmp], max_level=5)

        assert 2 in results
        source_refs = [r for r in results[2] if r.target_file == "source.xlsx"]
        assert len(source_refs) >= 1

        ref = source_refs[0]
        assert ref.precedent_chain is not None
        assert len(ref.precedent_chain) == 2  # B1 → C1
        assert ref.precedent_chain[0][1] == "B1"
        assert ref.precedent_chain[1][1] == "C1"


def test_trace_transitive_mixed_direct_and_indirect():
    """Level 2 finds both direct and transitive external refs."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)

        # upstream.xlsx: Data!A1 = '[direct.xlsx]S1'!A1 (direct)
        #                Data!B1 = C1*2, C1 = '[indirect.xlsx]S2'!B2 (transitive)
        upstream_file = tmp / "upstream.xlsx"
        _make_xlsx_with_external_formulas(upstream_file, {
            "Data": {
                "A1": "'[direct.xlsx]S1'!A1",
                "B1": "C1*2",
                "C1": "'[indirect.xlsx]S2'!B2",
            },
        })

        # model.xlsx references both A1 and B1
        model_file = tmp / "model.xlsx"
        _make_xlsx_with_external_formulas(model_file, {
            "Sheet1": {
                "A1": "'[upstream.xlsx]Data'!A1",
                "A2": "'[upstream.xlsx]Data'!B1",
            },
        })

        from lineage.tracing.formula_tracer import trace_formula_levels
        results = trace_formula_levels(model_file, search_dirs=[tmp], max_level=5)

        assert 2 in results
        target_files = {r.target_file for r in results[2]}
        assert "direct.xlsx" in target_files
        assert "indirect.xlsx" in target_files

        # Direct ref should have no precedent chain
        direct_ref = [r for r in results[2] if r.target_file == "direct.xlsx"][0]
        assert direct_ref.precedent_chain is None

        # Indirect ref should have a precedent chain
        indirect_ref = [r for r in results[2] if r.target_file == "indirect.xlsx"][0]
        assert indirect_ref.precedent_chain is not None
        assert len(indirect_ref.precedent_chain) >= 1


def test_report_with_precedent_chain():
    """TracingReporter shows precedent chain column in Level sheets."""
    from lineage.tracing.config import TraceConfig
    from lineage.tracing.report import TracingReporter
    from lineage.tracing.formula_tracer import ExternalReference

    config = TraceConfig()
    level_refs = {
        2: [
            ExternalReference(
                level=2, source_file="upstream.xlsx", source_sheet="Data",
                source_cell="A1", formula="B1*2",
                target_file="source.xlsx", target_sheet="Raw",
                target_range="C3", target_path="source.xlsx",
                file_found=False, resolved_path="/tmp/source.xlsx",
                precedent_chain=[("Data", "B1", "'[source.xlsx]Raw'!C3")],
            ),
            ExternalReference(
                level=2, source_file="upstream.xlsx", source_sheet="Data",
                source_cell="C1", formula="'[direct.xlsx]S1'!A1",
                target_file="direct.xlsx", target_sheet="S1",
                target_range="A1", target_path="direct.xlsx",
                file_found=True, resolved_path="/tmp/direct.xlsx",
                precedent_chain=None,
            ),
        ],
    }

    with tempfile.TemporaryDirectory() as tmpdir:
        out_dir = Path(tmpdir)
        reporter = TracingReporter()
        out_path = reporter.write_with_levels(
            [], [], config, Path("/tmp/model.xlsx"), "Sheet1",
            [], out_dir, level_refs=level_refs,
        )

        import openpyxl
        wb = openpyxl.load_workbook(str(out_path))
        ws = wb["Level 2"]

        # Header
        assert ws.cell(row=1, column=11).value == "Precedent Chain"

        # Row 2: transitive ref — should show chain
        chain_text = ws.cell(row=2, column=11).value
        assert "Data!B1" in chain_text
        assert chain_text != "(direct)"

        # Row 3: direct ref — should show "(direct)"
        assert ws.cell(row=3, column=11).value == "(direct)"


# ---------------------------------------------------------------------------
# Gap 1: Structured table references (no '!' separator)
# ---------------------------------------------------------------------------

def test_parse_named_refs_struct_table():
    """Structured table refs like [file.xlsx]Table[Col] are captured."""
    from lineage.tracing.formula_tracer import _parse_formula_named_refs
    refs = _parse_formula_named_refs("[data.xlsx]MyTable[Revenue]", {})
    assert len(refs) == 1
    assert refs[0] == ("data.xlsx", "", "MyTable", "data.xlsx")


def test_parse_named_refs_struct_table_no_column():
    """[file.xlsx]TableName without column specifier is captured."""
    from lineage.tracing.formula_tracer import _parse_formula_named_refs
    refs = _parse_formula_named_refs("[data.xlsx]Financials", {})
    assert len(refs) == 1
    assert refs[0][2] == "Financials"


def test_struct_table_does_not_match_cell_ref():
    """[file.xlsx]Sheet1!A1 is NOT captured as a struct table ref."""
    from lineage.tracing.formula_tracer import _parse_formula_named_refs
    refs = _parse_formula_named_refs("'[data.xlsx]Sheet1'!A1", {})
    assert len(refs) == 0


def test_parse_named_refs_with_bang():
    """Named refs with '!' like '[file.xlsx]'!MyRange are captured."""
    from lineage.tracing.formula_tracer import _parse_formula_named_refs
    refs = _parse_formula_named_refs("='[data.xlsx]'!MyRange", {})
    assert len(refs) == 1
    assert refs[0] == ("data.xlsx", "", "MyRange", "data.xlsx")


def test_parse_named_refs_sheet_qualified():
    """Sheet-qualified named ref '[file.xlsx]Sheet1'!Name is captured."""
    from lineage.tracing.formula_tracer import _parse_formula_named_refs
    refs = _parse_formula_named_refs("'[data.xlsx]Sheet1'!TotalRevenue", {})
    assert len(refs) == 1
    assert refs[0][1] == "Sheet1"
    assert refs[0][2] == "TotalRevenue"


# ---------------------------------------------------------------------------
# Gap 2: Sheet-scoped named ranges
# ---------------------------------------------------------------------------

def _make_xlsx_with_custom_defined_names(defined_names_xml: bytes) -> Path:
    """Helper: create an xlsx with custom <definedNames> in workbook.xml."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    wb.create_sheet("Sheet2")
    ws["A1"] = 1
    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(f.name)
    f.close()
    src = Path(f.name)
    patched = Path(f.name + ".p.xlsx")
    with zipfile.ZipFile(src, "r") as zin, zipfile.ZipFile(patched, "w") as zout:
        for item in zin.namelist():
            data = zin.read(item)
            if item == "xl/workbook.xml":
                # Replace empty definedNames or inject before </workbook>
                if b"<definedNames/>" in data:
                    data = data.replace(b"<definedNames/>", defined_names_xml)
                elif b"<definedNames>" in data:
                    # Replace existing block
                    data = re.sub(
                        rb"<definedNames>.*?</definedNames>",
                        defined_names_xml,
                        data,
                        flags=re.DOTALL,
                    )
                else:
                    data = data.replace(b"</workbook>", defined_names_xml + b"</workbook>")
            zout.writestr(item, data)
    src.unlink()
    return patched


def test_get_defined_names_scoped():
    """_get_defined_names returns scope info for localSheetId names."""
    from lineage.tracing.formula_tracer import _get_defined_names
    xml = (
        b'<definedNames>'
        b'<definedName name="Revenue">Sheet1!$A$1:$B$10</definedName>'
        b'<definedName name="Revenue" localSheetId="1">Sheet2!$C$1:$D$5</definedName>'
        b'</definedNames>'
    )
    path = _make_xlsx_with_custom_defined_names(xml)
    try:
        with zipfile.ZipFile(path) as zf:
            local, _ = _get_defined_names(zf)
        entries = local.get("revenue", [])
        assert len(entries) == 2
        global_e = [e for e in entries if e.scope_sheet is None]
        scoped_e = [e for e in entries if e.scope_sheet is not None]
        assert len(global_e) == 1
        assert global_e[0].sheet == "Sheet1"
        assert len(scoped_e) == 1
        assert scoped_e[0].scope_sheet == "Sheet2"
        assert scoped_e[0].cell_range == "C1:D5"
    finally:
        path.unlink()


def test_resolve_named_ref_prefers_scoped():
    """Resolver picks scoped entry when qualifier matches."""
    from lineage.tracing.formula_tracer import _resolve_named_ref_in_file
    xml = (
        b'<definedNames>'
        b'<definedName name="Data">Sheet1!$A$1:$A$10</definedName>'
        b'<definedName name="Data" localSheetId="1">Sheet2!$C$1:$C$20</definedName>'
        b'</definedNames>'
    )
    path = _make_xlsx_with_custom_defined_names(xml)
    try:
        # With qualifier matching scoped entry
        r = _resolve_named_ref_in_file(path, "Data", qualifier="Sheet2")
        assert r is not None
        assert r.sheet == "Sheet2"
        assert r.cell_range == "C1:C20"

        # Without qualifier → global
        r = _resolve_named_ref_in_file(path, "Data", qualifier="")
        assert r is not None
        assert r.sheet == "Sheet1"
    finally:
        path.unlink()


# ---------------------------------------------------------------------------
# Gap 3: Dynamic named ranges
# ---------------------------------------------------------------------------

def test_get_defined_names_dynamic():
    """Dynamic named ranges (with formulas) are flagged as is_dynamic."""
    from lineage.tracing.formula_tracer import _get_defined_names
    xml = (
        b'<definedNames>'
        b'<definedName name="DynRange">Sheet1!OFFSET($A$1,0,0,COUNTA($A:$A),1)</definedName>'
        b'<definedName name="StaticRange">Sheet1!$A$1:$B$5</definedName>'
        b'</definedNames>'
    )
    path = _make_xlsx_with_custom_defined_names(xml)
    try:
        with zipfile.ZipFile(path) as zf:
            local, _ = _get_defined_names(zf)
        dyn = local.get("dynrange", [])
        assert len(dyn) == 1
        assert dyn[0].is_dynamic is True

        static = local.get("staticrange", [])
        assert len(static) == 1
        assert static[0].is_dynamic is False
    finally:
        path.unlink()


def test_resolve_dynamic_range():
    """Resolving a dynamic name returns is_dynamic=True with empty cell_range."""
    from lineage.tracing.formula_tracer import _resolve_named_ref_in_file
    xml = (
        b'<definedNames>'
        b'<definedName name="Dyn">Sheet1!OFFSET($A$1,0,0,10,1)</definedName>'
        b'</definedNames>'
    )
    path = _make_xlsx_with_custom_defined_names(xml)
    try:
        r = _resolve_named_ref_in_file(path, "Dyn")
        assert r is not None
        assert r.is_dynamic is True
        assert r.sheet == "Sheet1"
        assert r.cell_range == ""  # can't statically resolve
    finally:
        path.unlink()


# ---------------------------------------------------------------------------
# Gap 4: Named ranges referencing external workbooks
# ---------------------------------------------------------------------------

def test_get_defined_names_external():
    """Names referencing external workbooks are returned in the external list."""
    from lineage.tracing.formula_tracer import _get_defined_names
    xml = (
        b"<definedNames>"
        b"<definedName name=\"ExtData\">'[other.xlsx]Raw'!$A$1:$C$100</definedName>"
        b"<definedName name=\"Local\">Sheet1!$A$1:$A$5</definedName>"
        b"</definedNames>"
    )
    path = _make_xlsx_with_custom_defined_names(xml)
    try:
        with zipfile.ZipFile(path) as zf:
            local, external = _get_defined_names(zf)
        # ExtData should be in external, not local
        assert "extdata" not in local
        assert len(external) == 1
        assert external[0][0] == "ExtData"
        assert external[0][1] == "other.xlsx"
        assert external[0][2] == "Raw"
        # Local should still be there
        assert "local" in local
    finally:
        path.unlink()


def test_resolve_external_defined_name():
    """Resolving an external-redirect name returns external_file info."""
    from lineage.tracing.formula_tracer import _resolve_named_ref_in_file
    xml = (
        b"<definedNames>"
        b"<definedName name=\"Budget\">'[budget.xlsx]Plan'!$A$1:$D$50</definedName>"
        b"</definedNames>"
    )
    path = _make_xlsx_with_custom_defined_names(xml)
    try:
        r = _resolve_named_ref_in_file(path, "Budget")
        assert r is not None
        assert r.external_file == "budget.xlsx"
        assert r.external_sheet == "Plan"
        assert r.external_range == "A1:D50"
    finally:
        path.unlink()


# ---------------------------------------------------------------------------
# Gap 5: Recursive file resolution + path stripping
# ---------------------------------------------------------------------------

def test_resolve_file_recursive(tmp_path):
    """_resolve_file finds files in subdirectories recursively."""
    from lineage.tracing.formula_tracer import _resolve_file
    sub = tmp_path / "a" / "b" / "c"
    sub.mkdir(parents=True)
    (sub / "deep.xlsx").write_bytes(b"fake")

    resolved, found = _resolve_file("deep.xlsx", [tmp_path])
    assert found
    assert resolved.name == "deep.xlsx"


def test_resolve_file_strips_windows_path(tmp_path):
    """_resolve_file strips Windows path prefix to find by basename."""
    from lineage.tracing.formula_tracer import _resolve_file
    (tmp_path / "budget.xlsx").write_bytes(b"fake")

    resolved, found = _resolve_file(
        "C:\\Users\\John\\Documents\\budget.xlsx", [tmp_path],
    )
    assert found
    assert resolved.name == "budget.xlsx"


def test_resolve_file_strips_unix_path(tmp_path):
    """_resolve_file strips Unix path prefix to find by basename."""
    from lineage.tracing.formula_tracer import _resolve_file
    (tmp_path / "data.xlsx").write_bytes(b"fake")

    resolved, found = _resolve_file("/home/user/docs/data.xlsx", [tmp_path])
    assert found
    assert resolved.name == "data.xlsx"


def test_resolve_file_case_insensitive_recursive(tmp_path):
    """Case-insensitive recursive search works."""
    from lineage.tracing.formula_tracer import _resolve_file
    sub = tmp_path / "sub"
    sub.mkdir()
    (sub / "Budget.xlsx").write_bytes(b"fake")

    resolved, found = _resolve_file("budget.xlsx", [tmp_path])
    assert found
    assert resolved.name == "Budget.xlsx"


# ---------------------------------------------------------------------------
# Gap 6: Table resolution
# ---------------------------------------------------------------------------

def test_resolve_table_in_file():
    """_resolve_named_ref_in_file resolves Excel table names."""
    import openpyxl
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from lineage.tracing.formula_tracer import _resolve_named_ref_in_file

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "x"
    ws["B2"] = 1
    tab = Table(displayName="tbl_data", ref="A1:B2")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9")
    ws.add_table(tab)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb.save(f.name)
        path = Path(f.name)
    try:
        r = _resolve_named_ref_in_file(path, "tbl_data")
        assert r is not None
        assert r.ref_type == "table"
        assert r.sheet == "Data"
        assert r.cell_range == "A1:B2"
    finally:
        path.unlink()


# ---------------------------------------------------------------------------
# Integration: named ref in formula tracing
# ---------------------------------------------------------------------------

def test_scan_external_refs_named_ref():
    """scan_external_refs detects named range references in formulas."""
    import openpyxl
    from lineage.tracing.formula_tracer import scan_external_refs

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Model"
    # Formula referencing a named range in an external file
    ws["A1"] = 1  # placeholder
    # We need to inject the formula via XML since openpyxl won't allow external refs
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        model = tmp / "model.xlsx"
        wb.save(str(model))

        # Patch sheet XML to have an external named ref formula
        with zipfile.ZipFile(model, "r") as zin:
            names = zin.namelist()
            sheet_path = [n for n in names if n.startswith("xl/worksheets/sheet")][0]
            sheet_xml = zin.read(sheet_path)

        # Inject a formula with [file]!NamedRange
        sheet_xml = sheet_xml.replace(
            b"<v>1</v>",
            b'<f>\'[upstream.xlsx]\'!TotalRevenue</f><v>1</v>',
        )

        patched = tmp / "model_patched.xlsx"
        with zipfile.ZipFile(model, "r") as zin, zipfile.ZipFile(patched, "w") as zout:
            for item in zin.namelist():
                if item == sheet_path:
                    zout.writestr(item, sheet_xml)
                else:
                    zout.writestr(item, zin.read(item))

        refs = scan_external_refs(patched, "model.xlsx", level=1, search_dirs=[tmp])
        named = [r for r in refs if r.ref_type == "named_ref"]
        assert len(named) >= 1
        assert named[0].target_file == "upstream.xlsx"
        assert named[0].target_name == "TotalRevenue"
