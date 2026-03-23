#!/usr/bin/env python3
"""Generate three synthetic Excel files with a formula chain for upstream tracing.

Chain: model.xlsx  --(formulas)-->  upstream_A.xlsx  --(formulas)-->  upstream_B.xlsx

upstream_B.xlsx  (the root source)
  - "RawData" sheet: raw numeric data (hardcoded values)

upstream_A.xlsx  (intermediate)
  - "Processed" sheet:
      * Some cells with formulas referencing upstream_B.xlsx (e.g., ='[upstream_B.xlsx]RawData'!A1 * 2)
      * Some hardcoded values

model.xlsx  (the model being analysed)
  - "Analysis" sheet:
      * Formulas referencing upstream_A.xlsx (e.g., ='[upstream_A.xlsx]Processed'!B2)
      * Hardcoded vectors copy-pasted from upstream_A.xlsx (for value-based matching)
"""

import openpyxl
from pathlib import Path

OUT = Path(__file__).parent


def make_upstream_b():
    """Root source file with raw data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RawData"

    # Header row
    ws["A1"] = "Quarter"
    ws["B1"] = "Revenue"
    ws["C1"] = "COGS"
    ws["D1"] = "Headcount"

    # 8 quarters of data (hardcoded)
    data = [
        ("Q1-2024", 1500, 600, 120),
        ("Q2-2024", 1620, 648, 125),
        ("Q3-2024", 1750, 700, 130),
        ("Q4-2024", 1890, 756, 135),
        ("Q1-2025", 2010, 804, 140),
        ("Q2-2025", 2180, 872, 148),
        ("Q3-2025", 2350, 940, 155),
        ("Q4-2025", 2500, 1000, 160),
    ]
    for i, (q, rev, cogs, hc) in enumerate(data, 2):
        ws.cell(i, 1, q)
        ws.cell(i, 2, rev)
        ws.cell(i, 3, cogs)
        ws.cell(i, 4, hc)

    # A second sheet with exchange rates
    ws2 = wb.create_sheet("FXRates")
    ws2["A1"] = "Quarter"
    ws2["B1"] = "USD/EUR"
    fx = [0.92, 0.91, 0.93, 0.90, 0.89, 0.88, 0.91, 0.90]
    for i, (q, rate) in enumerate(zip([r[0] for r in data], fx), 2):
        ws2.cell(i, 1, q)
        ws2.cell(i, 2, rate)

    path = OUT / "upstream_B.xlsx"
    wb.save(path)
    print(f"Created {path}")
    return path


def make_upstream_a():
    """Intermediate file that references upstream_B.xlsx via formulas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Processed"

    ws["A1"] = "Quarter"
    ws["B1"] = "Gross Profit"
    ws["C1"] = "Rev per Head"
    ws["D1"] = "Revenue (EUR)"

    # Rows 2-9: formulas referencing upstream_B.xlsx
    for i in range(2, 10):
        ws.cell(i, 1, f"Q{((i-2)%4)+1}-{2024 + (i-2)//4}")
        # B = Revenue - COGS  (references upstream_B)
        ws.cell(i, 2).value = f"='[upstream_B.xlsx]RawData'!B{i}-'[upstream_B.xlsx]RawData'!C{i}"
        # C = Revenue / Headcount  (references upstream_B)
        ws.cell(i, 3).value = f"='[upstream_B.xlsx]RawData'!B{i}/'[upstream_B.xlsx]RawData'!D{i}"
        # D = Revenue * FX rate (references TWO sheets in upstream_B)
        ws.cell(i, 4).value = f"='[upstream_B.xlsx]RawData'!B{i}*'[upstream_B.xlsx]FXRates'!B{i}"

    # Also have a "Summary" sheet with some hardcoded values (manually entered)
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Metric"
    ws2["B1"] = "Value"
    # These are hardcoded summary stats
    summary_vals = [
        ("Total Revenue", 15800),
        ("Total COGS", 6320),
        ("Avg Headcount", 139),
        ("Growth Rate", 0.085),
    ]
    for i, (label, val) in enumerate(summary_vals, 2):
        ws2.cell(i, 1, label)
        ws2.cell(i, 2, val)

    path = OUT / "upstream_A.xlsx"
    wb.save(path)
    print(f"Created {path}")
    return path


def make_model():
    """Model file that references upstream_A.xlsx and has copy-pasted values."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analysis"

    # --- Section 1: Formulas linking to upstream_A ---
    ws["A1"] = "Quarter"
    ws["B1"] = "Gross Profit (from upstream)"
    ws["C1"] = "Rev per Head (from upstream)"
    ws["D1"] = "Margin %"

    for i in range(2, 10):
        ws.cell(i, 1, f"Q{((i-2)%4)+1}-{2024 + (i-2)//4}")
        # Formula referencing upstream_A
        ws.cell(i, 2).value = f"='[upstream_A.xlsx]Processed'!B{i}"
        ws.cell(i, 3).value = f"='[upstream_A.xlsx]Processed'!C{i}"
        # A local formula using the pulled data
        ws.cell(i, 4).value = f"=B{i}/('[upstream_A.xlsx]Processed'!B{i}+'[upstream_A.xlsx]Processed'!C{i})"

    # --- Section 2: Hardcoded values (copy-pasted from upstream_B) ---
    # These are the EXACT revenue numbers from upstream_B.xlsx RawData!B2:B9
    ws["F1"] = "Pasted Revenue"
    revenue_pasted = [1500, 1620, 1750, 1890, 2010, 2180, 2350, 2500]
    for i, v in enumerate(revenue_pasted, 2):
        ws.cell(i, 6, v)  # Column F — hardcoded, no formula

    # --- Section 3: Hardcoded COGS (copy-pasted from upstream_B) ---
    ws["G1"] = "Pasted COGS"
    cogs_pasted = [600, 648, 700, 756, 804, 872, 940, 1000]
    for i, v in enumerate(cogs_pasted, 2):
        ws.cell(i, 7, v)  # Column G — hardcoded, no formula

    # --- Section 4: FX rates (copy-pasted from upstream_B FXRates) ---
    ws["H1"] = "Pasted FX"
    fx_pasted = [0.92, 0.91, 0.93, 0.90, 0.89, 0.88, 0.91, 0.90]
    for i, v in enumerate(fx_pasted, 2):
        ws.cell(i, 8, v)  # Column H — hardcoded, no formula

    path = OUT / "model.xlsx"
    wb.save(path)
    print(f"Created {path}")
    return path


if __name__ == "__main__":
    make_upstream_b()
    make_upstream_a()
    make_model()
    print("\nDone! Files created in", OUT)
    print("\nFormula chain:")
    print("  model.xlsx  -->  upstream_A.xlsx  -->  upstream_B.xlsx")
    print("\nHardcoded vectors in model.xlsx copied from upstream_B.xlsx:")
    print("  F2:F9 = Revenue, G2:G9 = COGS, H2:H9 = FX rates")
